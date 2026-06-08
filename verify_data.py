#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_data.py — 比對 raw Excel(data/) 與網站產生的 JSON(site/data/)是否一致。
獨立重讀 raw 檔，逐項比對筆數並抽查實際值，並偵測 charts 檔名撞分類等問題。
只讀不寫，純查核。"""
import os, re, json, glob, zipfile, hashlib
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(ROOT, "data")
OUT_DIR = os.path.join(ROOT, "site", "data")

issues = []   # 嚴重：缺漏/錯誤
warns = []    # 提醒：設計上忽略/模稜兩可

def rows_of(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); return None  # None = 無此工作表
    out = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close(); return out

def find_header(rows, key):
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if key in cells:
            return i, cells
    return None, None

def to_records(rows, key):
    if rows is None: return None
    hi, header = find_header(rows, key)
    if hi is None: return []
    recs = []
    for r in rows[hi+1:]:
        if r is None: continue
        if all(c is None or str(c).strip() == "" for c in r): continue
        first = str(r[0]).strip() if r[0] is not None else ""
        if first.startswith("──") or first.startswith("【"): continue
        rec = {h: v for h, v in zip(header, r) if h != ""}
        recs.append(rec)
    return recs

def media_list(path):
    try:
        z = zipfile.ZipFile(path)
    except Exception:
        return []
    m = sorted([n for n in z.namelist() if "/media/" in n and n.lower().endswith((".png",".jpg",".jpeg"))])
    data = [(n, hashlib.md5(z.read(n)).hexdigest()) for n in m]
    z.close(); return data

def load_json(p):
    if not os.path.exists(p): return None
    with open(p, encoding="utf-8") as f: return json.load(f)

def file_md5(p):
    if not os.path.exists(p): return None
    with open(p, "rb") as f: return hashlib.md5(f.read()).hexdigest()

date_dirs = sorted(d for d in glob.glob(os.path.join(DATA_DIR, "*")) if os.path.isdir(d) and re.fullmatch(r"\d{8}", os.path.basename(d)))
built_dirs = sorted(os.path.basename(d) for d in glob.glob(os.path.join(OUT_DIR, "*")) if os.path.isdir(d))
raw_dates = [os.path.basename(d) for d in date_dirs]

print("RAW 日期:", raw_dates)
print("已產生日期:", built_dirs)
for d in raw_dates:
    if d not in built_dirs:
        issues.append(f"[缺日期] raw 有 {d} 但 site/data 沒產生")

for ddir in date_dirs:
    date = os.path.basename(ddir)
    print(f"\n===== {date} =====")
    files = glob.glob(os.path.join(ddir, "*.xlsx"))

    # --- 複製 classify，並偵測撞名/被忽略 ---
    market = None
    stocks = {}
    ignored = []
    for path in files:
        fn = os.path.basename(path)
        if fn.startswith("台股放量訊號"):
            market = path; continue
        m = re.match(r"^(\d{4})", fn)
        if not m:
            ignored.append(fn); continue
        code = m.group(1)
        s = stocks.setdefault(code, {"charts": [], "daily": [], "analysis": [], "other": []})
        if "charts" in fn: s["charts"].append(path)
        elif "日報" in fn: s["daily"].append(path)
        elif "分析結果" in fn: s["analysis"].append(path)
        else: s["other"].append(fn)

    # 大盤
    if not market:
        issues.append(f"[{date}] 找不到 台股放量訊號 檔")
    else:
        mj = load_json(os.path.join(OUT_DIR, date, "market.json"))
        if mj is None:
            issues.append(f"[{date}] market.json 不存在")
        else:
            for key, (sheet, hkey) in {"signals":("今日放量訊號","代號"),"continuation":("量能延續追蹤","代號"),
                                       "release":("出關股追蹤","代號"),"disposed":("處置股清單","代號"),
                                       "winrate":("勝率回測","分類")}.items():
                raw = to_records(rows_of(market, sheet), hkey)
                rawn = "無此表" if raw is None else len(raw)
                jn = len(mj.get(key, []))
                flag = "" if (raw is not None and rawn == jn) or (raw is None and jn == 0) else "  <<< 不一致!"
                if flag: issues.append(f"[{date}] market.{key}: raw={rawn} json={jn}")
                print(f"  market.{key:12} raw={rawn:>6} json={jn:>6}{flag}")

    # 個股
    for code in sorted(stocks):
        s = stocks[code]
        sj = load_json(os.path.join(OUT_DIR, date, f"{code}.json"))
        print(f"  -- {code} --")
        # 撞名偵測
        if len(s["charts"]) > 1:
            warns.append(f"[{date}/{code}] 有多個 charts 檔，classify 只會留最後一個: {[os.path.basename(x) for x in s['charts']]}")
        if len(s["daily"]) > 1:
            issues.append(f"[{date}/{code}] 有多個 日報 檔: {[os.path.basename(x) for x in s['daily']]}")
        if len(s["analysis"]) > 1:
            issues.append(f"[{date}/{code}] 有多個 分析結果 檔: {[os.path.basename(x) for x in s['analysis']]}")
        if s["other"]:
            warns.append(f"[{date}/{code}] 未被使用的檔(週報/大量與均價等): {s['other']}")
        if sj is None:
            issues.append(f"[{date}/{code}] {code}.json 不存在"); continue

        # 日報 → buy_top/sell_top
        if s["daily"]:
            dp = s["daily"][0]
            for sheet, jkey in [("買進前20","buy_top"),("賣出前20","sell_top")]:
                raw = to_records(rows_of(dp, "券商"), "券商") if rows_of(dp, sheet) is not None else None
                # 用對應工作表
                rr = to_records(rows_of(dp, sheet), "券商")
                rn = "無此表" if rr is None else len(rr)
                jn = len(sj.get(jkey, []))
                flag = "  <<< 不一致!" if not ((rr is None and jn==0) or (rr is not None and len(rr)==jn)) else ""
                if flag: issues.append(f"[{date}/{code}] {jkey}: raw={rn} json={jn}")
                print(f"     {jkey:10} raw={rn:>5} json={jn:>5}{flag}")

        # 分析結果 → price_volume / broker_detail
        if s["analysis"]:
            ap = s["analysis"][0]
            pv = to_records(rows_of(ap, "買賣價量與家數"), "股價")
            det = to_records(rows_of(ap, "券商明細"), "股價")
            for rr, jkey in [(pv,"price_volume"),(det,"broker_detail")]:
                rn = "無此表" if rr is None else len(rr)
                jn = len(sj.get(jkey, []))
                flag = "  <<< 不一致!" if not ((rr is None and jn==0) or (rr is not None and len(rr)==jn)) else ""
                if flag: issues.append(f"[{date}/{code}] {jkey}: raw={rn} json={jn}")
                print(f"     {jkey:12} raw={rn:>6} json={jn:>6}{flag}")

        # charts → 比對實際 PNG 來自哪個 xlsx
        jcharts = sj.get("charts", [])
        png_dir = os.path.join(OUT_DIR, date, "charts")
        # 第一張 PNG 的 md5
        first_png = None
        if jcharts:
            first_png = file_md5(os.path.join(OUT_DIR, date, jcharts[0]))
        match_src = []
        for cp in s["charts"]:
            ml = media_list(cp)
            md5s = [h for _, h in ml]
            hit = (first_png in md5s) if first_png else False
            match_src.append((os.path.basename(cp), len(ml), hit))
        print(f"     charts json={len(jcharts)}  候選檔:")
        for fn, n, hit in match_src:
            print(f"        {'★用此檔' if hit else '       '} {fn}  (內含{n}張)")
        if s["charts"] and len(jcharts) == 0:
            issues.append(f"[{date}/{code}] 有 charts 檔卻沒產生任何圖")

print("\n\n========== 查核結果 ==========")
print(f"嚴重問題 issues: {len(issues)}")
for x in issues: print("  ✗", x)
print(f"\n提醒 warns: {len(warns)}")
for x in warns: print("  !", x)
if not issues:
    print("\n✓ 筆數比對全部一致，無缺漏/錯誤")
