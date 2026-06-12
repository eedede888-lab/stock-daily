"""
台股成交量分析系統 v3.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
使用方式（手動執行，建議每2~3日更新一次）：

  首次初始化（抓取7個月歷史）：
    python volume_system.py --init

  日常更新：
    python volume_system.py --update

  查詢個股：
    python volume_system.py --stock 2330

  查詢近期訊號：
    python volume_system.py --signals --days 30
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import sys, time, sqlite3, logging, requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

# ── 強制終端機以 UTF-8 輸出，避免在 Windows cp950 環境下列印 emoji 當機 ──
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ══════════════════════════════════════════════════════════════════════════════
# 全域設定
# ══════════════════════════════════════════════════════════════════════════════
DB_PATH = "tw_volume.db"
OUT_DIR = Path(r"C:\Users\羞羞的家\Desktop")
OUT_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
log = logging.getLogger(__name__)

# ── 放量閾值（五個維度）
THRESHOLDS = {
    "vs_prev":  {"★★★": 3.0, "★★": 2.0, "★": 1.5},
    "vs_ma5":   {"★★★": 2.5, "★★": 1.8, "★": 1.5},
    "vs_ma10":  {"★★★": 2.8, "★★": 1.9, "★": 1.5},
    "vs_ma20":  {"★★★": 3.0, "★★": 2.0, "★": 1.5},
    "vs_ma60":  {"★★★": 3.5, "★★": 2.5, "★": 2.0},
}

# ── 出關判斷設定
RELEASE_CFG = {
    "strong_ratio":   1.2,   # 出關量 >= 處置前量 × 1.2 → RELEASE_STRONG
    "normal_low":     0.8,   # 出關量 >= 處置前量 × 0.8 → RELEASE_NORMAL
    "surge_pct":      5.0,   # 量縮但漲幅 >= 5% → RELEASE_WEAK_SURGE
    "track_days":     5,     # 出關後追蹤天數
}

# ── 延續判斷設定
CONTINUE_CFG = {
    "strong_ratio":   1.5,   # 今日量 >= MA20 × 1.5 → CONTINUE_STRONG（改用MA20為基準）
    "normal_low":     1.0,   # 今日量 >= MA20 × 1.0 → CONTINUE_NORMAL
    "surge_pct":      5.0,   # 量縮但漲幅 >= 5% → CONTINUE_SURGE
}

# ── 訊號類型對照
SIGNAL_TYPES = {
    "NORMAL":              ("一般放量",       "一般"),
    "DISPOSED":            ("處置中放量",     "🔒 處置中"),
    "RELEASE_STRONG":      ("出關強勢",       "⚡ 出關強勢"),
    "RELEASE_NORMAL":      ("出關正常恢復",   "🔓 出關恢復"),
    "RELEASE_WEAK_SURGE":  ("出關量縮價強",   "🔥 出關價強"),
    "RELEASE_HOLD_MA":     ("出關守均線",     "📊 出關守均"),
    "CONTINUE_STRONG":     ("次日強力延續",   "🚀 強力延續"),
    "CONTINUE_NORMAL":     ("次日正常延續",   "✅ 正常延續"),
    "CONTINUE_SURGE":      ("次日價強量縮",   "🔥 價強量縮"),
}

# ── 處置期間分析常數
LIMIT_UP_PCT   =  9.5   # 漲停門檻 %
LIMIT_DOWN_PCT = -9.5   # 跌停門檻 %
SURGE_DAY_PCT  =  6.0   # 單日強漲門檻 %


# ══════════════════════════════════════════════════════════════════════════════
# 資料庫初始化
# ══════════════════════════════════════════════════════════════════════════════
def init_db(db_path: str = DB_PATH) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    con.executescript("""
        CREATE TABLE IF NOT EXISTS daily_volume (
            stock_id    TEXT    NOT NULL,
            date        TEXT    NOT NULL,
            volume      INTEGER NOT NULL,
            close_price REAL,
            change_pct  REAL,
            market      TEXT,
            is_disposed INTEGER DEFAULT 0,
            PRIMARY KEY (stock_id, date)
        );

        CREATE TABLE IF NOT EXISTS volume_ma (
            stock_id   TEXT NOT NULL,
            date       TEXT NOT NULL,
            vol_prev   INTEGER,
            close_prev REAL,
            ma5        REAL,
            ma10       REAL,
            ma20       REAL,
            ma60       REAL,
            PRIMARY KEY (stock_id, date)
        );

        CREATE TABLE IF NOT EXISTS signals (
            stock_id       TEXT NOT NULL,
            date           TEXT NOT NULL,
            market         TEXT,
            signal_type    TEXT,
            signal_star    TEXT,
            signal_detail  TEXT,
            volume         INTEGER,
            close_price    REAL,
            change_pct     REAL,
            vol_price_rel  TEXT,
            vol_prev       INTEGER,
            close_prev     REAL,
            ma5            REAL,
            ma10           REAL,
            ma20           REAL,
            ma60           REAL,
            ratio_prev     REAL,
            ratio_ma5      REAL,
            ratio_ma10     REAL,
            ratio_ma20     REAL,
            ratio_ma60     REAL,
            is_disposed    INTEGER DEFAULT 0,
            PRIMARY KEY (stock_id, date)
        );

        CREATE TABLE IF NOT EXISTS disposed_stocks (
            stock_id      TEXT NOT NULL,
            fetch_date    TEXT NOT NULL,
            dispose_start TEXT,
            dispose_end   TEXT,
            reason        TEXT,
            market        TEXT,
            PRIMARY KEY (stock_id, fetch_date)
        );

        CREATE TABLE IF NOT EXISTS release_tracking (
            stock_id            TEXT NOT NULL,
            release_date        TEXT NOT NULL,
            market              TEXT,
            release_type        TEXT,
            dispose_days        INTEGER,
            pre_disp_date       TEXT,
            pre_disp_volume     INTEGER,
            pre_disp_close      REAL,
            release_volume      INTEGER,
            release_close       REAL,
            release_change_pct  REAL,
            vol_ratio_vs_pre    REAL,
            price_change_vs_pre REAL,
            signal_star         TEXT,
            signal_detail       TEXT,
            disposal_period_chg  REAL,
            disposal_max_gain    REAL,
            disposal_max_loss    REAL,
            disposal_surge_count INTEGER,
            disposal_limit_up    INTEGER,
            disposal_limit_down  INTEGER,
            disposal_flag        TEXT,
            extra_note           TEXT,
            PRIMARY KEY (stock_id, release_date)
        );

        CREATE TABLE IF NOT EXISTS continuation_tracking (
            stock_id            TEXT NOT NULL,
            signal_date         TEXT NOT NULL,
            continue_date       TEXT NOT NULL,
            market              TEXT,
            continue_type       TEXT,
            signal_type         TEXT,
            signal_star         TEXT,
            signal_volume       INTEGER,
            signal_close        REAL,
            continue_volume     INTEGER,
            continue_close      REAL,
            continue_change_pct REAL,
            vol_ratio           REAL,
            cumulative_change   REAL,
            PRIMARY KEY (stock_id, signal_date, continue_date)
        );

        CREATE TABLE IF NOT EXISTS stock_info (
            stock_id     TEXT PRIMARY KEY,
            stock_name   TEXT,
            market       TEXT,
            updated_date TEXT
        );

        CREATE TABLE IF NOT EXISTS signal_tracking (
            stock_id          TEXT NOT NULL,
            signal_date       TEXT NOT NULL,
            signal_close      REAL,
            signal_star       TEXT,
            market            TEXT,
            stage             TEXT NOT NULL,
            stage_start_date  TEXT,
            stage_start_close REAL,
            status            TEXT DEFAULT 'tracking',
            PRIMARY KEY (stock_id, signal_date, stage)
        );

        CREATE TABLE IF NOT EXISTS signal_tracking_days (
            stock_id    TEXT NOT NULL,
            signal_date TEXT NOT NULL,
            stage       TEXT NOT NULL,
            day_num     INTEGER NOT NULL,
            trade_date  TEXT,
            close_price REAL,
            change_pct  REAL,
            volume      INTEGER,
            PRIMARY KEY (stock_id, signal_date, stage, day_num)
        );
    """)
    con.commit()
    migrate_db(con)
    return con


def migrate_db(con: sqlite3.Connection):
    """為既有資料庫補齊新版欄位（ALTER TABLE）"""
    new_cols = [
        ("release_tracking", "disposal_period_chg",  "REAL"),
        ("release_tracking", "disposal_max_gain",     "REAL"),
        ("release_tracking", "disposal_max_loss",     "REAL"),
        ("release_tracking", "disposal_surge_count",  "INTEGER"),
        ("release_tracking", "disposal_limit_up",     "INTEGER"),
        ("release_tracking", "disposal_limit_down",   "INTEGER"),
        ("release_tracking", "disposal_flag",         "TEXT"),
        ("release_tracking", "extra_note",            "TEXT"),
    ]
    for table, col, dtype in new_cols:
        try:
            con.execute(f"ALTER TABLE {table} ADD COLUMN {col} {dtype}")
        except Exception:
            pass

    # 補建 market_stats 表（若舊 DB 尚未建立）
    con.executescript("""
        CREATE TABLE IF NOT EXISTS market_stats (
            date           TEXT PRIMARY KEY,
            twse_avg_chg   REAL,
            twse_up_cnt    INTEGER,
            twse_down_cnt  INTEGER,
            twse_flat_cnt  INTEGER,
            twse_total     INTEGER,
            tpex_avg_chg   REAL,
            tpex_up_cnt    INTEGER,
            tpex_down_cnt  INTEGER,
            tpex_flat_cnt  INTEGER,
            tpex_total     INTEGER,
            market_flag    TEXT
        );
    """)
    con.commit()

    # 補建 signal_tracking 系列表（若舊 DB 尚未建立）
    con.executescript("""
        CREATE TABLE IF NOT EXISTS signal_tracking (
            stock_id          TEXT NOT NULL,
            signal_date       TEXT NOT NULL,
            signal_close      REAL,
            signal_star       TEXT,
            market            TEXT,
            stage             TEXT NOT NULL,
            stage_start_date  TEXT,
            stage_start_close REAL,
            status            TEXT DEFAULT 'tracking',
            PRIMARY KEY (stock_id, signal_date, stage)
        );
        CREATE TABLE IF NOT EXISTS signal_tracking_days (
            stock_id    TEXT NOT NULL,
            signal_date TEXT NOT NULL,
            stage       TEXT NOT NULL,
            day_num     INTEGER NOT NULL,
            trade_date  TEXT,
            close_price REAL,
            change_pct  REAL,
            volume      INTEGER,
            PRIMARY KEY (stock_id, signal_date, stage, day_num)
        );
        CREATE TABLE IF NOT EXISTS market_stats (
            date           TEXT PRIMARY KEY,
            twse_avg_chg   REAL,
            twse_up_cnt    INTEGER,
            twse_down_cnt  INTEGER,
            twse_flat_cnt  INTEGER,
            twse_total     INTEGER,
            tpex_avg_chg   REAL,
            tpex_up_cnt    INTEGER,
            tpex_down_cnt  INTEGER,
            tpex_flat_cnt  INTEGER,
            tpex_total     INTEGER,
            market_flag    TEXT
        );
    """)
    con.commit()


# ══════════════════════════════════════════════════════════════════════════════
# 資料抓取
# ══════════════════════════════════════════════════════════════════════════════
def _parse_num(val, default=0.0):
    """安全解析數字，移除逗號與空白"""
    try:
        return float(str(val).replace(",", "").replace("--", "0").strip())
    except Exception:
        return default


def fetch_twse_day(date_str: str) -> pd.DataFrame:
    """抓取 TWSE 全市場當日成交量＋收盤價，date_str='YYYYMMDD'"""
    url = (
        "https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX"
        f"?response=json&date={date_str}&type=ALL"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        js = r.json()
        if js.get("stat") != "OK":
            return pd.DataFrame()

        target = next(
            (t for t in js.get("tables", [])
             if "成交股數" in t.get("fields", []) and "證券代號" in t.get("fields", [])),
            None
        )
        if target is None:
            return pd.DataFrame()

        fields = target["fields"]
        df = pd.DataFrame(target["data"], columns=fields)
        df["stock_id"] = df["證券代號"].str.strip()

        # 只保留4位數純數字代號
        df = df[df["stock_id"].str.match(r"^\d{4}$")].copy()

        df["volume"] = df["成交股數"].apply(lambda x: int(_parse_num(x)))

        # 收盤價
        if "收盤價" in fields:
            df["close_price"] = df["收盤價"].apply(_parse_num)
        else:
            df["close_price"] = 0.0

        # 漲跌幅
        if "漲跌(+/-)" in fields and "漲跌價差" in fields:
            def calc_chg(row):
                try:
                    sign_raw = re.sub(r'<[^>]+>', '', str(row["漲跌(+/-)"])).strip()
                    sign  = 1.0 if sign_raw == "+" else (
                            -1.0 if sign_raw == "-" else 0.0)
                    diff  = _parse_num(row["漲跌價差"])
                    close = _parse_num(row["收盤價"])
                    prev  = close - sign * diff
                    return round(sign * diff / prev * 100, 2) if prev > 0 else 0.0
                except Exception:
                    return 0.0
            df["change_pct"] = df.apply(calc_chg, axis=1)
        else:
            df["change_pct"] = 0.0

        df["date"]   = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        df["market"] = "TWSE"
        df["stock_name"] = df["證券名稱"].str.strip() if "證券名稱" in df.columns else ""
        df = df[df["volume"] > 0]
        return df[["stock_id", "stock_name", "date", "volume", "close_price", "change_pct", "market"]].reset_index(drop=True)

    except Exception as e:
        log.warning(f"TWSE {date_str} 抓取失敗：{e}")
        return pd.DataFrame()


def fetch_tpex_day(date_str: str) -> pd.DataFrame:
    """抓取 TPEX 全市場當日成交量＋收盤價，date_str='YYYYMMDD'"""
    dt       = datetime.strptime(date_str, "%Y%m%d")
    roc      = dt.year - 1911
    date_roc = f"{roc}/{dt.strftime('%m/%d')}"
    url = (
        "https://www.tpex.org.tw/web/stock/aftertrading/"
        "otc_quotes_no1430/stk_wn1430_result.php"
        f"?l=zh-tw&d={date_roc}&se=EW&o=json"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        js     = r.json()
        tables = js.get("tables", [])
        if not tables or not tables[0].get("data"):
            return pd.DataFrame()

        tbl    = tables[0]
        fields = tbl["fields"]
        df     = pd.DataFrame(tbl["data"], columns=fields)

        code_col   = fields[0]   # 代號
        close_col  = fields[2]   # 收盤
        change_col = fields[3]   # 漲跌
        vol_col    = fields[7]   # 成交股數

        df["stock_id"] = df[code_col].str.strip()
        df = df[df["stock_id"].str.match(r"^\d{4}$")].copy()

        df["volume"] = df[vol_col].apply(lambda x: int(_parse_num(x)))

        df["close_price"] = df[close_col].apply(_parse_num)

        def calc_chg_tpex(row):
            try:
                diff  = _parse_num(row[change_col])
                close = _parse_num(row[close_col])
                prev  = close - diff
                return round(diff / prev * 100, 2) if prev > 0 else 0.0
            except Exception:
                return 0.0

        df["change_pct"] = df.apply(calc_chg_tpex, axis=1)
        df["date"]       = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        df["market"]     = "TPEX"
        name_col         = fields[1] if len(fields) > 1 else None
        df["stock_name"] = df[name_col].str.strip() if name_col and name_col in df.columns else ""
        df = df[df["volume"] > 0]
        return df[["stock_id", "stock_name", "date", "volume", "close_price", "change_pct", "market"]].reset_index(drop=True)

    except Exception as e:
        log.warning(f"TPEX {date_str} 抓取失敗：{e}")
        return pd.DataFrame()

# ══════════════════════════════════════════════════════════════════════════════
# 股票名稱工具
# ══════════════════════════════════════════════════════════════════════════════
def save_stock_names(con: sqlite3.Connection, name_df: pd.DataFrame):
    """將股票名稱 upsert 至 stock_info 表"""
    if name_df.empty:
        return
    for _, row in name_df.iterrows():
        try:
            con.execute("""
                INSERT OR REPLACE INTO stock_info (stock_id, stock_name, market, updated_date)
                VALUES (?, ?, ?, ?)
            """, (row["stock_id"], row.get("stock_name", ""),
                  row.get("market", ""), row.get("updated_date", "")))
        except Exception:
            pass
    con.commit()


def get_stock_name_map(con: sqlite3.Connection) -> dict:
    """回傳 {stock_id: stock_name} dict"""
    try:
        df = pd.read_sql("SELECT stock_id, stock_name FROM stock_info", con)
        return dict(zip(df["stock_id"], df["stock_name"].fillna("")))
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# 處置期間分析
# ══════════════════════════════════════════════════════════════════════════════
def analyze_disposal_period(
    con: sqlite3.Connection,
    stock_id: str,
    dispose_start: str,
    dispose_end: str,
) -> dict:
    """
    分析股票在處置期間內的股價行為。
    回傳 dict，包含：期間漲跌幅、最大單日漲跌、漲跌停次數、特殊標記。
    """
    if not dispose_start or not dispose_end:
        return {}

    df = pd.read_sql("""
        SELECT date, close_price, change_pct
        FROM daily_volume
        WHERE stock_id = ? AND date >= ? AND date <= ?
        ORDER BY date ASC
    """, con, params=(stock_id, dispose_start, dispose_end))

    if df.empty or len(df) < 1:
        return {}

    first_close = df.iloc[0]["close_price"]
    last_close  = df.iloc[-1]["close_price"]
    period_chg  = (
        round((last_close - first_close) / first_close * 100, 2)
        if first_close and first_close > 0 else None
    )

    gains = df["change_pct"].dropna()
    max_gain       = round(float(gains.max()), 2) if not gains.empty else None
    max_loss       = round(float(gains.min()), 2) if not gains.empty else None
    surge_count    = int((gains >= SURGE_DAY_PCT).sum())
    limit_up_cnt   = int((gains >= LIMIT_UP_PCT).sum())
    limit_down_cnt = int((gains <= LIMIT_DOWN_PCT).sum())

    flags = []
    if limit_up_cnt > 0:
        flags.append("處置期漲停")
    if limit_down_cnt > 0:
        flags.append("處置期跌停")
    if period_chg is not None and period_chg >= 15:
        flags.append("期間強漲")
    if period_chg is not None and period_chg <= -15:
        flags.append("期間大跌")

    return {
        "disposal_period_chg":  period_chg,
        "disposal_max_gain":    max_gain,
        "disposal_max_loss":    max_loss,
        "disposal_surge_count": surge_count,
        "disposal_limit_up":    limit_up_cnt,
        "disposal_limit_down":  limit_down_cnt,
        "disposal_flag":        " / ".join(flags) if flags else "",
    }


# ══════════════════════════════════════════════════════════════════════════════
# 訊號追蹤（5日/10日/15日/15日）
# ══════════════════════════════════════════════════════════════════════════════
STAGE_DAYS   = {"A": 5, "B": 10, "C": 15, "D": 15}
EXTEND_PCT   = 6.0   # 延伸門檻 %

def _get_trading_dates(con: sqlite3.Connection, from_date: str, n: int) -> list:
    """取 from_date 之後（含）的 n 個交易日"""
    rows = con.execute("""
        SELECT DISTINCT date FROM daily_volume
        WHERE date >= ? ORDER BY date LIMIT ?
    """, (from_date, n)).fetchall()
    return [r[0] for r in rows]


def _fill_stage_days(con, stock_id, signal_date, stage, start_date, max_days):
    """將 start_date 後 max_days 個交易日的日線資料寫入 signal_tracking_days"""
    dates = _get_trading_dates(con, start_date, max_days + 1)
    # 第0筆是 start_date 本身，跳過
    track_dates = dates[1:max_days + 1]
    for i, td in enumerate(track_dates, 1):
        row = con.execute("""
            SELECT close_price, change_pct, volume FROM daily_volume
            WHERE stock_id=? AND date=?
        """, (stock_id, td)).fetchone()
        if row:
            con.execute("""
                INSERT OR REPLACE INTO signal_tracking_days
                    (stock_id, signal_date, stage, day_num, trade_date,
                     close_price, change_pct, volume)
                VALUES (?,?,?,?,?,?,?,?)
            """, (stock_id, signal_date, stage, i, td, row[0], row[1], row[2]))


def _chg_pct(base_close, curr_close):
    if base_close and base_close > 0 and curr_close:
        return round((curr_close - base_close) / base_close * 100, 2)
    return None


def compute_signal_tracking(con: sqlite3.Connection, target_date: str):
    """
    每日更新 signal_tracking：
    1. 新增 Stage A 給今日出現的訊號
    2. 填入所有追蹤中的各 stage 當日資料
    3. 判斷是否達延伸條件，晉升下一 stage
    """
    # ── Step 0：強制結案過期追蹤（冷門股/資料缺失導致假追蹤中）────────────────
    # 修正：改用「實際交易日數」作為過期判斷依據，避免長假（如農曆春節）
    # 誤殺正常追蹤中的 Stage（改用 daily_volume 中最近 N 個交易日的最早日期作 cutoff）
    STALE_THRESHOLDS = {
        "A": int(STAGE_DAYS["A"] * 2) + 1,   # 5日 → 取前11個交易日的最早日
        "B": int(STAGE_DAYS["B"] * 2) + 1,   # 10日 → 取前21個交易日
        "C": int(STAGE_DAYS["C"] * 2) + 1,   # 15日 → 取前31個交易日
        "D": int(STAGE_DAYS["D"] * 2) + 1,   # 15日 → 取前31個交易日
    }
    for stage, n_trade_days in STALE_THRESHOLDS.items():
        # 取 target_date 之前（含）第 n_trade_days 個實際交易日
        cutoff_row = con.execute("""
            SELECT date FROM daily_volume
            WHERE date <= ? ORDER BY date DESC LIMIT 1 OFFSET ?
        """, (target_date, n_trade_days - 1)).fetchone()
        if cutoff_row:
            con.execute("""
                UPDATE signal_tracking
                SET status = 'closed'
                WHERE stage = ? AND status = 'tracking'
                  AND stage_start_date <= ?
            """, (stage, cutoff_row[0]))
    con.commit()

    # ── Step 1：今日新訊號 → 建立 Stage A ────────────────────────────────────
    new_sigs = con.execute("""
        SELECT stock_id, date, close_price, signal_star, market
        FROM signals
        WHERE date=? AND signal_type IN ('NORMAL','DISPOSED')
          AND stock_id NOT IN (
              SELECT stock_id FROM signal_tracking WHERE signal_date=? AND stage='A'
          )
    """, (target_date, target_date)).fetchall()

    for row in new_sigs:
        sid, sig_date, sig_close, star, market = row
        con.execute("""
            INSERT OR IGNORE INTO signal_tracking
                (stock_id, signal_date, signal_close, signal_star, market,
                 stage, stage_start_date, stage_start_close, status)
            VALUES (?,?,?,?,?, 'A',?,?,'tracking')
        """, (sid, sig_date, sig_close, star, market, sig_date, sig_close))

    con.commit()

    # ── Step 2：補填各 stage 當日資料 ────────────────────────────────────────
    tracking = con.execute("""
        SELECT stock_id, signal_date, stage, stage_start_date, stage_start_close,
               signal_close, signal_star, market
        FROM signal_tracking
        WHERE status='tracking'
    """).fetchall()

    for (sid, sig_date, stage, stg_start, stg_close,
         sig_close, star, market) in tracking:

        max_days = STAGE_DAYS[stage]

        # 計算目前已過幾個交易日
        existing = con.execute("""
            SELECT COUNT(*) FROM signal_tracking_days
            WHERE stock_id=? AND signal_date=? AND stage=?
        """, (sid, sig_date, stage)).fetchone()[0]

        if existing >= max_days:
            continue   # 本 stage 資料已滿，等待晉升判斷

        # 取 stg_start 之後第 existing+1 個交易日
        dates = _get_trading_dates(con, stg_start, max_days + 1)
        track_dates = dates[1:max_days + 1]

        if len(track_dates) <= existing:
            continue   # 尚無新交易日資料

        # 只補尚未填入的日期
        for i, td in enumerate(track_dates[existing:], existing + 1):
            row = con.execute("""
                SELECT close_price, change_pct, volume FROM daily_volume
                WHERE stock_id=? AND date=?
            """, (sid, td)).fetchone()
            if row:
                con.execute("""
                    INSERT OR REPLACE INTO signal_tracking_days
                        (stock_id, signal_date, stage, day_num, trade_date,
                         close_price, change_pct, volume)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (sid, sig_date, stage, i, td, row[0], row[1], row[2]))

    con.commit()

    # ── Step 3：判斷晉升 ────────────────────────────────────────────────────
    NEXT = {"A": "B", "B": "C", "C": "D"}

    for (sid, sig_date, stage, stg_start, stg_close,
         sig_close, star, market) in tracking:

        if stage not in NEXT:
            # Stage D 無下一階段，但需要主動結案（否則永遠卡在 tracking）
            max_days_d = STAGE_DAYS["D"]
            days_count_d = con.execute("""
                SELECT COUNT(*) FROM signal_tracking_days
                WHERE stock_id=? AND signal_date=? AND stage='D'
            """, (sid, sig_date)).fetchone()[0]
            if days_count_d >= max_days_d:
                con.execute("""
                    UPDATE signal_tracking SET status='closed'
                    WHERE stock_id=? AND signal_date=? AND stage='D'
                """, (sid, sig_date))
            continue   # Stage D 無下一階段

        max_days = STAGE_DAYS[stage]

        # 確認本 stage 資料已滿
        days_data = con.execute("""
            SELECT day_num, trade_date, close_price FROM signal_tracking_days
            WHERE stock_id=? AND signal_date=? AND stage=?
            ORDER BY day_num
        """, (sid, sig_date, stage)).fetchall()

        if len(days_data) < max_days:
            continue   # 尚未填滿，不判斷

        last_date  = days_data[-1][1]
        last_close = days_data[-1][2]

        if last_close is None:
            continue

        # 計算各基準漲跌幅
        chg_from_signal = _chg_pct(sig_close, last_close)

        # B段：只看訊號日
        if stage == "A":
            qualify = (chg_from_signal is not None
                       and abs(chg_from_signal) >= EXTEND_PCT)

        # C段：看D+5（stg_close）或訊號日
        elif stage == "B":
            chg_from_stage = _chg_pct(stg_close, last_close)
            qualify = (
                (chg_from_stage is not None and abs(chg_from_stage) >= EXTEND_PCT)
                or (chg_from_signal is not None and abs(chg_from_signal) >= EXTEND_PCT)
            )

        # D段：看D+15（stg_close）或D+5或訊號日
        else:  # stage == "C"
            chg_from_stage = _chg_pct(stg_close, last_close)
            # 取 B 段的 stage_start_close 作為 D+5 基準
            b_row = con.execute("""
                SELECT stage_start_close FROM signal_tracking
                WHERE stock_id=? AND signal_date=? AND stage='B'
            """, (sid, sig_date)).fetchone()
            chg_from_b = _chg_pct(b_row[0], last_close) if b_row else None
            qualify = (
                (chg_from_stage is not None and abs(chg_from_stage) >= EXTEND_PCT)
                or (chg_from_b is not None and abs(chg_from_b) >= EXTEND_PCT)
                or (chg_from_signal is not None and abs(chg_from_signal) >= EXTEND_PCT)
            )

        if qualify:
            next_stage = NEXT[stage]
            con.execute("""
                UPDATE signal_tracking SET status='promoted'
                WHERE stock_id=? AND signal_date=? AND stage=?
            """, (sid, sig_date, stage))
            con.execute("""
                INSERT OR IGNORE INTO signal_tracking
                    (stock_id, signal_date, signal_close, signal_star, market,
                     stage, stage_start_date, stage_start_close, status)
                VALUES (?,?,?,?,?, ?,?,?,'tracking')
            """, (sid, sig_date, sig_close, star, market,
                  next_stage, last_date, last_close))
        else:
            con.execute("""
                UPDATE signal_tracking SET status='closed'
                WHERE stock_id=? AND signal_date=? AND stage=?
            """, (sid, sig_date, stage))

    con.commit()
    log.info(f"signal_tracking 更新完成（{target_date}）")


def backfill_signal_tracking(con: sqlite3.Connection, start_date: str = "2026-04-06"):
    """
    從 start_date 開始回溯補寫所有歷史追蹤資料。
    按時間順序逐日執行 compute_signal_tracking。
    """
    log.info(f"開始回溯 signal_tracking（從 {start_date}）...")

    all_dates = [r[0] for r in con.execute("""
        SELECT DISTINCT date FROM signals
        WHERE date >= ? ORDER BY date
    """, (start_date,)).fetchall()]

    log.info(f"  共 {len(all_dates)} 個交易日需要回溯")

    for i, date in enumerate(all_dates, 1):
        compute_signal_tracking(con, date)
        if i % 10 == 0 or i == len(all_dates):
            log.info(f"  [{i}/{len(all_dates)}] {date} 完成")

    total = con.execute("SELECT COUNT(*) FROM signal_tracking").fetchone()[0]
    log.info(f"回溯完成：signal_tracking 共 {total} 筆")


def fetch_disposed_stocks() -> pd.DataFrame:
    """抓取 TWSE + TPEX 處置股清單"""
    today   = datetime.today().strftime("%Y-%m-%d")
    results = []

    def _roc_to_ad(roc_str: str) -> str:
        """民國日期 '115/04/17' → '2026-04-17'"""
        roc_str = roc_str.strip()
        parts   = roc_str.replace("-", "/").split("/")
        if len(parts) == 3:
            try:
                y = int(parts[0]) + 1911
                return f"{y}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            except Exception:
                pass
        return ""

    def _parse_dispose_api(js: dict, market: str):
        """通用解析 TWSE / TPEX 處置股 JSON"""
        # 支援兩種結構：最外層直接有 fields/data，或包在 tables 裡
        if "tables" in js:
            tables = js["tables"]
            if not tables:
                return
            fields = tables[0].get("fields", [])
            data   = tables[0].get("data",   [])
        else:
            fields = js.get("fields", [])
            data   = js.get("data",   [])

        # 找欄位索引（相容不同欄位名稱）
        def _idx(candidates):
            for c in candidates:
                if c in fields:
                    return fields.index(c)
            return None

        idx_code  = _idx(["證券代號"])
        idx_range = _idx(["處置起迄時間", "處置起訖時間"])
        idx_reason= _idx(["處置條件", "處置原因"])

        if idx_code is None or idx_range is None:
            log.warning(f"{market} 找不到必要欄位，fields={fields}")
            return

        for row in data:
            code = str(row[idx_code]).strip()
            # 只保留4位純數字（排除權證、ETF等）
            if not code.isdigit() or len(code) != 4:
                continue

            # 解析起迄時間，格式：'115/04/17～115/04/30' 或 '115/04/30~115/05/14'
            raw_range  = str(row[idx_range]).strip()
            date_parts = raw_range.replace("～", "~").split("~")
            d_start    = _roc_to_ad(date_parts[0].strip()) if len(date_parts) >= 1 else ""
            d_end      = _roc_to_ad(date_parts[1].strip()) if len(date_parts) >= 2 else ""

            reason = str(row[idx_reason]) if idx_reason is not None else ""

            results.append({
                "stock_id":      code,
                "fetch_date":    today,
                "dispose_start": d_start,
                "dispose_end":   d_end,
                "reason":        reason,
                "market":        market,
            })

    # ── TWSE ─────────────────────────────────────────────────────────────────
    try:
        r  = requests.get(
            "https://www.twse.com.tw/rwd/zh/announcement/punish?response=json",
            timeout=15
        )
        r.raise_for_status()
        _parse_dispose_api(r.json(), "TWSE")
        log.info(f"TWSE 處置股：{sum(1 for x in results if x['market']=='TWSE')} 筆")
    except Exception as e:
        log.warning(f"TWSE 處置股抓取失敗：{e}")

    # ── TPEX ─────────────────────────────────────────────────────────────────
    before = len(results)
    try:
        r  = requests.get(
            "https://www.tpex.org.tw/web/bulletin/disposal_information/"
            "disposal_information_result.php?l=zh-tw&o=json",
            timeout=15
        )
        r.raise_for_status()
        _parse_dispose_api(r.json(), "TPEX")
        log.info(f"TPEX 處置股：{len(results) - before} 筆")
    except Exception as e:
        log.warning(f"TPEX 處置股抓取失敗：{e}")

    cols = ["stock_id", "fetch_date", "dispose_start", "dispose_end", "reason", "market"]
    return pd.DataFrame(results, columns=cols) if results else pd.DataFrame(columns=cols)

# ══════════════════════════════════════════════════════════════════════════════
# 歷史批次下載
# ══════════════════════════════════════════════════════════════════════════════
def _trading_dates(start: datetime, end: datetime) -> list:
    dates, cur = [], start
    while cur <= end:
        if cur.weekday() < 5:
            dates.append(cur.strftime("%Y%m%d"))
        cur += timedelta(days=1)
    return dates


def fix_zero_change_pct(con: sqlite3.Connection, date_str: str):
    """
    修正 daily_volume 中 change_pct=0 的股票。
    原因：TWSE 除息/除權日 漲跌(+/-) 回傳 'X'，導致 sign=0，change_pct 被誤算為 0。
    修正方式：用昨日實際收盤價重新計算真實漲跌幅。
    只修正 close_price 與前日不同的股票（真平盤不動）。
    """
    prev_row = con.execute("""
        SELECT date FROM daily_volume
        WHERE date < ? ORDER BY date DESC LIMIT 1
    """, (date_str,)).fetchone()
    if not prev_row:
        return
    prev_date = prev_row[0]

    rows = con.execute("""
        SELECT t.stock_id, t.close_price, p.close_price AS prev_close
        FROM daily_volume t
        JOIN daily_volume p ON t.stock_id = p.stock_id AND p.date = ?
        WHERE t.date = ? AND t.change_pct = 0 AND p.close_price > 0
    """, (prev_date, date_str)).fetchall()

    fixed = 0
    for sid, close, prev_close in rows:
        if close is None or prev_close is None or prev_close == 0:
            continue
        if abs(close - prev_close) < 1e-8:
            continue   # 真平盤，不修正
        actual_chg = round((close - prev_close) / prev_close * 100, 2)
        con.execute(
            "UPDATE daily_volume SET change_pct=? WHERE stock_id=? AND date=?",
            (actual_chg, sid, date_str)
        )
        fixed += 1

    if fixed:
        con.commit()
        log.info(f"漲跌幅補正（除息/除權）：{fixed} 筆（{date_str}）")


def fetch_history(con: sqlite3.Connection, months_back: int = 7):
    end       = datetime.today()
    start     = end - timedelta(days=30 * months_back)
    raw_dates = _trading_dates(start, end)

    existing = {r[0] for r in con.execute("SELECT DISTINCT date FROM daily_volume").fetchall()}
    fmt      = lambda d: f"{d[:4]}-{d[4:6]}-{d[6:]}"
    todo     = [(d, fmt(d)) for d in raw_dates if fmt(d) not in existing]

    log.info(f"需補抓 {len(todo)} 個交易日（共 {len(raw_dates)} 日）")
    for i, (raw, fmted) in enumerate(todo):
        log.info(f"[{i+1}/{len(todo)}] {fmted} ...")
        twse_df = fetch_twse_day(raw)
        tpex_df = fetch_tpex_day(raw)
        valid   = [df for df in [twse_df, tpex_df] if not df.empty]
        if not valid:
            log.info(f"  {fmted} 無資料（非交易日），跳過")
            continue
        rows = pd.concat(valid, ignore_index=True)
        rows = rows[rows["volume"] > 0]

        # 儲存股票名稱（在去重複前執行，確保名稱被記錄）
        if "stock_name" in rows.columns and not rows.empty:
            name_df = rows[["stock_id", "stock_name", "market"]].copy()
            name_df["updated_date"] = fmted
            save_stock_names(con, name_df)
            rows = rows.drop(columns=["stock_name"])

        if not rows.empty:
            existing_ids = {
                r[0] for r in con.execute(
                    "SELECT stock_id FROM daily_volume WHERE date=?", (fmted,)
                ).fetchall()
            }
            rows = rows[~rows["stock_id"].isin(existing_ids)]
            if not rows.empty:
                rows.to_sql("daily_volume", con, if_exists="append", index=False)
                con.commit()
                # 補正除息/除權造成的漲跌幅誤算
                fix_zero_change_pct(con, fmted)
        # 計算該日大盤情境統計
        compute_market_stats(con, fmted)
        time.sleep(0.6)
    log.info("歷史資料補抓完成")


# ══════════════════════════════════════════════════════════════════════════════
# 處置標記
# ══════════════════════════════════════════════════════════════════════════════
def mark_disposed_volume(con: sqlite3.Connection):
    """
    依照 disposed_stocks 表的 dispose_start / dispose_end，
    將 daily_volume 對應日期的 is_disposed 標記為 1。
    """
    log.info("標記處置期間成交量...")
    disposed = pd.read_sql(
        "SELECT DISTINCT stock_id, dispose_start, dispose_end FROM disposed_stocks",
        con
    )
    count = 0
    for _, row in disposed.iterrows():
        sid   = row["stock_id"]
        start = str(row.get("dispose_start", "") or "")
        end   = str(row.get("dispose_end",   "") or "")
        if not start:
            continue
        # 若 dispose_end 為空，代表仍在處置中，使用今天
        if not end:
            end = datetime.today().strftime("%Y-%m-%d")
        # 日期格式統一（可能是民國年或西元年）
        start = _normalize_date(start)
        end   = _normalize_date(end)
        if not start:
            continue
        cur = con.execute("""
            UPDATE daily_volume SET is_disposed = 1
            WHERE stock_id = ? AND date >= ? AND date <= ?
        """, (sid, start, end))
        count += cur.rowcount
    con.commit()
    log.info(f"處置標記完成，共更新 {count} 筆")


def _normalize_date(date_str: str) -> str:
    """嘗試解析各種日期格式，回傳 'YYYY-MM-DD'"""
    date_str = str(date_str).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # 嘗試民國年 (113-01-01)
    try:
        parts = date_str.split("-")
        if len(parts) == 3 and len(parts[0]) <= 3:
            y = int(parts[0]) + 1911
            return f"{y}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    except Exception:
        pass
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# 移動平均計算（含 MA10）
# ══════════════════════════════════════════════════════════════════════════════
def compute_moving_averages(con: sqlite3.Connection):
    log.info("計算 MA5 / MA10 / MA20 / MA60 及前日量、前日收盤...")
    df = pd.read_sql(
        "SELECT stock_id, date, volume, close_price FROM daily_volume ORDER BY stock_id, date",
        con
    )
    if df.empty:
        log.warning("daily_volume 無資料")
        return

    df["date"] = pd.to_datetime(df["date"])
    df.sort_values(["stock_id", "date"], inplace=True)

    g = df.groupby("stock_id")
    df["vol_prev"]   = g["volume"].transform(lambda x: x.shift(1))
    df["close_prev"] = g["close_price"].transform(lambda x: x.shift(1))
    df["ma5"]        = g["volume"].transform(lambda x: x.rolling(5,  min_periods=3).mean())
    df["ma10"]       = g["volume"].transform(lambda x: x.rolling(10, min_periods=5).mean())
    df["ma20"]       = g["volume"].transform(lambda x: x.rolling(20, min_periods=10).mean())
    df["ma60"]       = g["volume"].transform(lambda x: x.rolling(60, min_periods=20).mean())

    df["date"] = df["date"].dt.strftime("%Y-%m-%d")
    ma_df = df[["stock_id", "date", "vol_prev", "close_prev", "ma5", "ma10", "ma20", "ma60"]]

    con.execute("DELETE FROM volume_ma")
    ma_df.to_sql("volume_ma", con, if_exists="append", index=False)
    con.commit()
    log.info(f"均量計算完成，{len(ma_df):,} 筆")


# ══════════════════════════════════════════════════════════════════════════════
# 訊號偵測工具函式
# ══════════════════════════════════════════════════════════════════════════════
def _star_level(ratio, key: str) -> str:
    if ratio is None or pd.isna(ratio):
        return ""
    t = THRESHOLDS[key]
    if ratio >= t["★★★"]: return "★★★"
    if ratio >= t["★★"]:  return "★★"
    if ratio >= t["★"]:   return "★"
    return ""


def _build_signal(row) -> tuple:
    """回傳 (最高星等, 詳細說明字串)"""
    parts, stars = [], []

    checks = [
        ("ratio_prev",  "vs_prev",  "較前日",  "倍"),
        ("ratio_ma5",   "vs_ma5",   "較MA5×",  ""),
        ("ratio_ma10",  "vs_ma10",  "較MA10×", ""),
        ("ratio_ma20",  "vs_ma20",  "較MA20×", ""),
        ("ratio_ma60",  "vs_ma60",  "較MA60×", ""),
    ]
    for col, tkey, label, suffix in checks:
        val = row.get(col)
        if val is None or pd.isna(val) or val <= 0:
            continue
        s = _star_level(val, tkey)
        if s:
            parts.append(f"{label}{val:.1f}{suffix}{s}")
            stars.append(s)

    if not parts:
        return "", ""

    rank     = {"★★★": 3, "★★": 2, "★": 1}
    top_star = max(stars, key=lambda x: rank[x])
    return top_star, " | ".join(parts)


def _vol_price_rel(volume, ma5, change_pct) -> str:
    """量價關係判斷（含平盤）"""
    try:
        vol_up = float(volume) >= float(ma5) if ma5 and not pd.isna(ma5) else False
        chg    = float(change_pct) if change_pct is not None and not pd.isna(change_pct) else None
        if chg is None:
            return ""
        if chg == 0:
            return "平盤"
        prc_up = chg > 0
        if vol_up  and prc_up:      return "量增價漲 ↑↑"
        if vol_up  and not prc_up:  return "量增價跌 ↑↓"
        if not vol_up and prc_up:   return "量縮價漲 ↓↑"
        return "量縮價跌 ↓↓"
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# 主要訊號偵測（NORMAL / DISPOSED）
# ══════════════════════════════════════════════════════════════════════════════
def detect_signals(
    con: sqlite3.Connection,
    target_date: str,
    disposed_ids: set,
) -> pd.DataFrame:
    log.info(f"偵測一般放量訊號（{target_date}）...")
    df = pd.read_sql("""
        SELECT v.stock_id, v.date, v.volume, v.close_price, v.change_pct,
               v.market, v.is_disposed,
               m.vol_prev, m.close_prev, m.ma5, m.ma10, m.ma20, m.ma60
        FROM daily_volume v
        JOIN volume_ma m ON v.stock_id = m.stock_id AND v.date = m.date
        WHERE v.date = ?
    """, con, params=(target_date,))

    if df.empty:
        log.warning(f"{target_date} 無資料")
        return df

    for col in ["ratio_prev", "ratio_ma5", "ratio_ma10", "ratio_ma20", "ratio_ma60"]:
        df[col] = None

    df["ratio_prev"]  = df.apply(lambda r: r["volume"]/r["vol_prev"]
                                  if r["vol_prev"] and r["vol_prev"]>0 else None, axis=1)
    df["ratio_ma5"]   = df.apply(lambda r: r["volume"]/r["ma5"]
                                  if r["ma5"] and r["ma5"]>0 else None, axis=1)
    df["ratio_ma10"]  = df.apply(lambda r: r["volume"]/r["ma10"]
                                  if r["ma10"] and r["ma10"]>0 else None, axis=1)
    df["ratio_ma20"]  = df.apply(lambda r: r["volume"]/r["ma20"]
                                  if r["ma20"] and r["ma20"]>0 else None, axis=1)
    df["ratio_ma60"]  = df.apply(lambda r: r["volume"]/r["ma60"]
                                  if r["ma60"] and r["ma60"]>0 else None, axis=1)

    sig = df.apply(_build_signal, axis=1, result_type="expand")
    sig.columns = ["signal_star", "signal_detail"]
    df = pd.concat([df, sig], axis=1)
    df = df[df["signal_star"] != ""].copy()

    df["signal_type"] = df.apply(
        lambda r: "DISPOSED" if (r["is_disposed"] or r["stock_id"] in disposed_ids)
                  else "NORMAL",
        axis=1
    )
    df["vol_price_rel"] = df.apply(
        lambda r: _vol_price_rel(r["volume"], r["ma5"], r["change_pct"]), axis=1
    )
    df["is_disposed"] = df["stock_id"].apply(lambda x: 1 if x in disposed_ids else 0)

    # 寫入資料庫
    save_cols = [
        "stock_id", "date", "market", "signal_type", "signal_star", "signal_detail",
        "volume", "close_price", "change_pct", "vol_price_rel",
        "vol_prev", "close_prev", "ma5", "ma10", "ma20", "ma60",
        "ratio_prev", "ratio_ma5", "ratio_ma10", "ratio_ma20", "ratio_ma60", "is_disposed"
    ]
    if not df.empty:
        df[save_cols].to_sql("signals", con, if_exists="append", index=False)
        con.commit()

    log.info(f"一般訊號：{len(df)} 檔（"
             f"★★★:{(df['signal_star']=='★★★').sum()} | "
             f"★★:{(df['signal_star']=='★★').sum()} | "
             f"★:{(df['signal_star']=='★').sum()} | "
             f"處置中:{(df['signal_type']=='DISPOSED').sum()} | "
             f"量<{200}張已過濾）")
    # ── 情境過濾欄位（方向二）──────────────────────────────────────────────────
    # 取得大盤情境
    mkt_ctx = get_market_context(con, target_date)
    df["market_chg_pct"] = df["market"].map({
        "TWSE": mkt_ctx.get("twse_avg_chg"),
        "TPEX": mkt_ctx.get("tpex_avg_chg"),
    })
    df["market_flag"] = mkt_ctx.get("market_flag", "")

    # 重複訊號標記（同標的 10 日內已有訊號）
    df["repeat_flag"] = df["stock_id"].apply(
        lambda sid: "⚠️重複" if _check_repeat_signal(con, sid, target_date) else ""
    )

    # 情境標籤彙整（供 Excel 顯示）
    def _ctx_label(row):
        flags = []
        mflag = row["market_flag"]
        if mflag:
            emoji = next((e for _, _, n, e in MARKET_LEVELS if n == mflag), "")
            if mflag not in ("中性", ""):
                flags.append(f"{emoji}{mflag}")
        if row["repeat_flag"]:
            flags.append(row["repeat_flag"])
        return " ".join(flags)

    df["context"] = df.apply(_ctx_label, axis=1)

    # ── 排序：星等 DESC → 成交金額 DESC，並過濾極低流動性標的 ──────────────────
    MIN_VOLUME = 200
    df = df[df["volume"] >= MIN_VOLUME].copy()
    df["turnover"] = df["close_price"].fillna(0) * df["volume"].fillna(0)
    df = df.sort_values(
        ["signal_star", "turnover"],
        ascending=[False, False],
        key=lambda c: c.map({"★★★": 3, "★★": 2, "★": 1})
                      if c.name == "signal_star" else c
    ).drop(columns=["turnover"])
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 出關訊號偵測
# ══════════════════════════════════════════════════════════════════════════════
def _get_pre_disposition_baseline(con: sqlite3.Connection, stock_id: str, dispose_start: str):
    """取處置開始前最後一個交易日的量與收盤價"""
    row = con.execute("""
        SELECT volume, close_price, date FROM daily_volume
        WHERE stock_id = ? AND date < ? AND is_disposed = 0
        ORDER BY date DESC LIMIT 1
    """, (stock_id, dispose_start)).fetchone()
    if row:
        return {"volume": row[0], "close": row[1], "date": row[2]}
    return None


def detect_release_signals(
    con: sqlite3.Connection,
    target_date: str,
    disposed_ids: set,
) -> pd.DataFrame:
    """偵測出關後第一個交易日的訊號"""
    log.info(f"偵測出關訊號（{target_date}）...")

    # 找出最近 30 天內有處置紀錄，但今日不在處置名單的股票
    cutoff = (datetime.strptime(target_date, "%Y-%m-%d") - timedelta(days=30)).strftime("%Y-%m-%d")
    recently_disposed = set(
        r[0] for r in con.execute("""
            SELECT DISTINCT stock_id FROM disposed_stocks
            WHERE dispose_start >= ? OR dispose_end >= ?
        """, (cutoff, cutoff)).fetchall()
    )
    candidates = recently_disposed - disposed_ids
    if not candidates:
        log.info("無出關候選標的")
        return pd.DataFrame()

    results = []
    for sid in candidates:
        # 確認今日有交易資料
        today_row = con.execute("""
                    SELECT v.volume, v.close_price, v.change_pct, v.market, m.ma5, m.ma10, m.ma20, m.ma60
                    FROM daily_volume v
                    JOIN volume_ma m ON v.stock_id=m.stock_id AND v.date=m.date
                    WHERE v.stock_id=? AND v.date=?
                """, (sid, target_date)).fetchone()
        if not today_row:
            continue

        vol, close, chg_pct, market, ma5, ma10, ma20, ma60 = today_row

        # 找處置紀錄
        disp_row = con.execute("""
            SELECT dispose_start, dispose_end FROM disposed_stocks
            WHERE stock_id=? ORDER BY fetch_date DESC LIMIT 1
        """, (sid,)).fetchone()
        if not disp_row:
            continue

        d_start = _normalize_date(str(disp_row[0] or ""))
        d_end   = _normalize_date(str(disp_row[1] or ""))
        if not d_start:
            continue

        # 確認今日是出關後前 N 個交易日
        try:
            end_dt    = datetime.strptime(d_end, "%Y-%m-%d") if d_end else datetime.today()
            today_dt  = datetime.strptime(target_date, "%Y-%m-%d")
            days_diff = (today_dt - end_dt).days
            if days_diff < 0 or days_diff > RELEASE_CFG["track_days"] * 2:
                continue
        except Exception:
            continue

        # 取處置天數
        try:
            start_dt  = datetime.strptime(d_start, "%Y-%m-%d")
            disp_days = (end_dt - start_dt).days
        except Exception:
            disp_days = 0

        # 取處置前基準
        baseline = _get_pre_disposition_baseline(con, sid, d_start)
        if not baseline:
            continue

        pre_vol   = baseline["volume"]
        pre_close = baseline["close"]
        pre_date  = baseline["date"]

        if pre_vol <= 0:
            continue

        vol_ratio         = vol / pre_vol
        price_change_pre  = round((close - pre_close) / pre_close * 100, 2) if pre_close > 0 else 0.0

        # ── 處置期間分析 ─────────────────────────────────────────────────────
        dinfo = analyze_disposal_period(con, sid, d_start, d_end)

        # ── 判斷出關類型 ─────────────────────────────────────────────────────
        if vol_ratio >= RELEASE_CFG["strong_ratio"]:
            rtype = "RELEASE_STRONG"
        elif vol_ratio >= RELEASE_CFG["normal_low"]:
            rtype = "RELEASE_NORMAL"
        elif chg_pct is not None and chg_pct >= RELEASE_CFG["surge_pct"]:
            # 只偵測上漲，修正原本 abs() 的錯誤
            rtype = "RELEASE_WEAK_SURGE"
        elif (close and ma20 and ma20 > 0
              and close > ma20
              and vol_ratio >= 0.3):
            # 量未達標，但收盤仍在 MA20 上方 → 結構未破壞，值得觀察
            rtype = "RELEASE_HOLD_MA"
        else:
            continue  # 量縮且無任何強勢訊號，排除

        # ── 備註邏輯 ─────────────────────────────────────────────────────────
        extra_note = ""
        if rtype == "RELEASE_HOLD_MA" and dinfo.get("disposal_limit_down", 0) > 0:
            extra_note = "⚠️ 跌停後守均（疑似洗盤）"
        elif dinfo.get("disposal_limit_up", 0) > 0:
            extra_note = "🔥 處置期出現漲停"
        elif dinfo.get("disposal_period_chg") is not None and dinfo["disposal_period_chg"] >= 15:
            extra_note = "📈 處置期間強漲"

        # ── 計算放量訊號強度 ──────────────────────────────────────────────────
        ratio_ma20 = vol / ma20 if ma20 and ma20 > 0 else None
        ratio_ma60 = vol / ma60 if ma60 and ma60 > 0 else None
        star, detail = _build_signal({
            "ratio_prev": vol_ratio,
            "ratio_ma5":  vol/ma5  if ma5  and ma5  > 0 else None,
            "ratio_ma10": vol/ma10 if ma10 and ma10 > 0 else None,
            "ratio_ma20": ratio_ma20,
            "ratio_ma60": ratio_ma60,
        })

        results.append({
            "stock_id":            sid,
            "release_date":        target_date,
            "market":              market,
            "release_type":        rtype,
            "dispose_days":        disp_days,
            "pre_disp_date":       pre_date,
            "pre_disp_volume":     pre_vol,
            "pre_disp_close":      pre_close,
            "release_volume":      vol,
            "release_close":       close,
            "release_change_pct":  chg_pct,
            "vol_ratio_vs_pre":    round(vol_ratio, 2),
            "price_change_vs_pre": price_change_pre,
            "signal_star":         star,
            "signal_detail":       detail,
            # 處置期間分析欄位
            "disposal_period_chg":  dinfo.get("disposal_period_chg"),
            "disposal_max_gain":    dinfo.get("disposal_max_gain"),
            "disposal_max_loss":    dinfo.get("disposal_max_loss"),
            "disposal_surge_count": dinfo.get("disposal_surge_count", 0),
            "disposal_limit_up":    dinfo.get("disposal_limit_up",    0),
            "disposal_limit_down":  dinfo.get("disposal_limit_down",  0),
            "disposal_flag":        dinfo.get("disposal_flag", ""),
            "extra_note":           extra_note,
        })

    if not results:
        log.info("無出關訊號")
        return pd.DataFrame()

    df = pd.DataFrame(results)
    df.to_sql("release_tracking", con, if_exists="append", index=False)
    con.commit()
    log.info(f"出關訊號：{len(df)} 檔（強勢:{(df['release_type']=='RELEASE_STRONG').sum()} | "
             f"正常:{(df['release_type']=='RELEASE_NORMAL').sum()} | "
             f"量縮價強:{(df['release_type']=='RELEASE_WEAK_SURGE').sum()} | "
             f"守均:{(df['release_type']=='RELEASE_HOLD_MA').sum()}）")
    return df.sort_values("vol_ratio_vs_pre", ascending=False)


# ══════════════════════════════════════════════════════════════════════════════
# 量能延續追蹤
# ══════════════════════════════════════════════════════════════════════════════
def detect_continuation(
    con: sqlite3.Connection,
    target_date: str,
) -> pd.DataFrame:
    """
    找出前 N 日內的放量訊號，檢查今日量能是否延續。
    vol_ratio 改用「今日量 / 今日MA20」取代「今日量 / 前日訊號量」，
    避免前日放量峰值導致比值偏低、大量漏接的問題。
    """
    log.info(f"偵測量能延續（{target_date}）...")

    # 前一交易日
    prev_date_row = con.execute("""
        SELECT MAX(date) FROM daily_volume WHERE date < ?
    """, (target_date,)).fetchone()
    if not prev_date_row or not prev_date_row[0]:
        return pd.DataFrame()
    prev_date = prev_date_row[0]

    # 取前日所有訊號（NORMAL 或 CONTINUE 皆可作為延續基準）
    # 修正：CONTINUE 的 signal_close 改用原始訊號日收盤（signals.close_price），
    # 而非前日延續的收盤（continue_close），確保 cumulative_change 反映相較訊號日的真實累計漲幅
    prev_signals = pd.read_sql("""
        SELECT s.stock_id, s.date, s.market, s.signal_type, s.signal_star,
               s.volume AS signal_volume, s.close_price AS signal_close
        FROM signals s
        WHERE s.date = ?
        UNION
        SELECT c.stock_id, c.signal_date AS date, c.market,
               c.continue_type AS signal_type, c.signal_star,
               c.continue_volume AS signal_volume,
               COALESCE(orig.close_price, c.signal_close) AS signal_close
        FROM continuation_tracking c
        LEFT JOIN signals orig
            ON orig.stock_id = c.stock_id AND orig.date = c.signal_date
        WHERE c.continue_date = ?
    """, con, params=(prev_date, prev_date))

    if prev_signals.empty:
        log.info("前日無訊號，跳過延續偵測")
        return pd.DataFrame()

    # 取今日量價資料，同時帶入 MA20 作為比較基準
    today_data = pd.read_sql("""
        SELECT d.stock_id,
               d.volume      AS cont_vol,
               d.close_price AS cont_close,
               d.change_pct,
               m.ma20        AS cont_ma20
        FROM daily_volume d
        JOIN volume_ma m ON d.stock_id = m.stock_id AND d.date = m.date
        WHERE d.date = ?
          AND m.ma20 > 0
    """, con, params=(target_date,))

    if today_data.empty:
        return pd.DataFrame()

    merged = prev_signals.merge(today_data, on="stock_id", how="inner")
    if merged.empty:
        return pd.DataFrame()

    # ── 核心修正：vol_ratio 改為今日量 / MA20 ────────────────────────────────
    # 舊邏輯：cont_vol / signal_volume（前日放量高峰，分母偏大 → 大量漏接）
    # 新邏輯：cont_vol / MA20（長期均量基準，門檻一致且穩定）
    merged["vol_ratio"] = (merged["cont_vol"] / merged["cont_ma20"]).round(3)

    def classify(row):
        r   = row["vol_ratio"]
        chg = row.get("change_pct") or 0.0
        if r >= CONTINUE_CFG["strong_ratio"]:        # ≥ 1.5 × MA20
            return "CONTINUE_STRONG"
        elif r >= CONTINUE_CFG["normal_low"]:        # ≥ 1.0 × MA20
            return "CONTINUE_NORMAL"
        elif chg >= CONTINUE_CFG["surge_pct"]:       # 量縮但價格仍強勢上漲
            return "CONTINUE_SURGE"
        return ""

    merged["continue_type"] = merged.apply(classify, axis=1)
    merged = merged[merged["continue_type"] != ""].copy()

    if merged.empty:
        log.info("無量能延續標的")
        return pd.DataFrame()

    merged["cumulative_change"] = merged.apply(
        lambda r: round((r["cont_close"] - r["signal_close"]) / r["signal_close"] * 100, 2)
        if r["signal_close"] and r["signal_close"] > 0 else 0.0,
        axis=1
    )

    # 避免重複寫入（同一標的同日可能從 signals 與 continuation_tracking 各帶出一筆）
    merged = merged.drop_duplicates(subset=["stock_id"], keep="first")

    out = pd.DataFrame({
        "stock_id":            merged["stock_id"],
        "signal_date":         merged["date"],
        "continue_date":       target_date,
        "market":              merged["market"],
        "continue_type":       merged["continue_type"],
        "signal_type":         merged["signal_type"],
        "signal_star":         merged["signal_star"],
        "signal_volume":       merged["signal_volume"],
        "signal_close":        merged["signal_close"],
        "continue_volume":     merged["cont_vol"],
        "continue_close":      merged["cont_close"],
        "continue_change_pct": merged["change_pct"],
        "vol_ratio":           merged["vol_ratio"],
        "cumulative_change":   merged["cumulative_change"],
    })

    # 避免 DB 重複寫入
    existing = {(r[0], r[1]) for r in con.execute(
        "SELECT stock_id, continue_date FROM continuation_tracking WHERE continue_date=?",
        (target_date,)
    ).fetchall()}
    out = out[~out.apply(lambda r: (r["stock_id"], r["continue_date"]) in existing, axis=1)]

    if not out.empty:
        out.to_sql("continuation_tracking", con, if_exists="append", index=False)
        con.commit()

    log.info(f"量能延續：{len(out)} 檔（"
             f"強力:{(out['continue_type']=='CONTINUE_STRONG').sum()} | "
             f"正常:{(out['continue_type']=='CONTINUE_NORMAL').sum()} | "
             f"價強:{(out['continue_type']=='CONTINUE_SURGE').sum()}）")
    return out.sort_values("vol_ratio", ascending=False)



# ══════════════════════════════════════════════════════════════════════════════
# 情境過濾模組（方向二）
# ══════════════════════════════════════════════════════════════════════════════

# 大盤旗標門檻設定
# 大盤市況十一段分級門檻（TWSE全市場個股簡單平均漲跌幅，[lo, hi) 邊界規則）
MARKET_LEVELS = [
    (-999,  -2.5, '極跌', '🔻'),   # ≤ -2.5%
    (-2.5,  -1.8, '強跌', '🌧️'),  # -2.5% ~ -1.8%
    (-1.8,  -1.0, '大跌', '📉'),   # -1.8% ~ -1.0%
    (-1.0,  -0.3, '小跌', '🔽'),   # -1.0% ~ -0.3%
    (-0.3,  -0.1, '偏跌', '↘️'),  # -0.3% ~ -0.1%
    (-0.1,   0.1, '中性', '➡️'),  # -0.1% ~ +0.1%
    ( 0.1,   0.3, '偏漲', '↗️'),  # +0.1% ~ +0.3%
    ( 0.3,   1.0, '小漲', '🔼'),   # +0.3% ~ +1.0%
    ( 1.0,   1.8, '大漲', '📈'),   # +1.0% ~ +1.8%
    ( 1.8,   2.5, '強漲', '🌊'),   # +1.8% ~ +2.5%
    ( 2.5,   999, '極漲', '🚀'),   # ≥ +2.5%
]

def _classify_market(chg: float) -> tuple:
    """回傳 (級別名稱, emoji)"""
    for lo, hi, name, emoji in MARKET_LEVELS:
        if lo <= chg < hi:
            return name, emoji
    return '極漲', '🚀'  # fallback
REPEAT_SIGNAL_DAYS =  3     # N 個交易日內同標的重複訊號門檻（用實際交易日，非曆法天）


def compute_market_stats(con: sqlite3.Connection, date_str: str) -> None:
    """
    從 daily_volume 計算當日大盤統計，寫入 market_stats。
    以全市場平均漲跌幅作為大盤強弱的代理指標。
    date_str: 'YYYY-MM-DD'
    """
    # 各市場分別統計
    stats: dict = {"date": date_str}
    for mkt, prefix in [("TWSE", "twse"), ("TPEX", "tpex")]:
        rows = con.execute("""
            SELECT change_pct FROM daily_volume
            WHERE date=? AND market=? AND change_pct IS NOT NULL
              AND volume > 0 AND is_disposed = 0
        """, (date_str, mkt)).fetchall()

        chg_list = [r[0] for r in rows if r[0] is not None]
        if not chg_list:
            stats[f"{prefix}_avg_chg"]  = None
            stats[f"{prefix}_up_cnt"]   = None
            stats[f"{prefix}_down_cnt"] = None
            stats[f"{prefix}_flat_cnt"] = None
            stats[f"{prefix}_total"]    = None
        else:
            stats[f"{prefix}_avg_chg"]  = round(sum(chg_list) / len(chg_list), 3)
            stats[f"{prefix}_up_cnt"]   = sum(1 for c in chg_list if c > 0)
            stats[f"{prefix}_down_cnt"] = sum(1 for c in chg_list if c < 0)
            stats[f"{prefix}_flat_cnt"] = sum(1 for c in chg_list if c == 0)
            stats[f"{prefix}_total"]    = len(chg_list)

    # 大盤旗標（以 TWSE 為主，十一段分級）
    twse_chg = stats.get("twse_avg_chg")
    if twse_chg is None:
        stats["market_flag"] = ""
    else:
        level_name, _ = _classify_market(twse_chg)
        stats["market_flag"] = level_name

    con.execute("""
        INSERT OR REPLACE INTO market_stats
        (date, twse_avg_chg, twse_up_cnt, twse_down_cnt, twse_flat_cnt, twse_total,
               tpex_avg_chg,  tpex_up_cnt,  tpex_down_cnt,  tpex_flat_cnt,  tpex_total,
               market_flag)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        stats["date"],
        stats.get("twse_avg_chg"), stats.get("twse_up_cnt"),
        stats.get("twse_down_cnt"), stats.get("twse_flat_cnt"), stats.get("twse_total"),
        stats.get("tpex_avg_chg"), stats.get("tpex_up_cnt"),
        stats.get("tpex_down_cnt"), stats.get("tpex_flat_cnt"), stats.get("tpex_total"),
        stats["market_flag"],
    ))
    con.commit()


def get_market_context(con: sqlite3.Connection, date_str: str) -> dict:
    """
    取得指定日期的大盤情境資料。
    若尚未計算，自動觸發 compute_market_stats。
    """
    row = con.execute(
        "SELECT twse_avg_chg, tpex_avg_chg, market_flag FROM market_stats WHERE date=?",
        (date_str,)
    ).fetchone()

    if row is None:
        compute_market_stats(con, date_str)
        row = con.execute(
            "SELECT twse_avg_chg, tpex_avg_chg, market_flag FROM market_stats WHERE date=?",
            (date_str,)
        ).fetchone()

    return {
        "twse_avg_chg": row[0] if row else None,
        "tpex_avg_chg": row[1] if row else None,
        "market_flag":  row[2] if row else "",
    }


def _check_repeat_signal(con: sqlite3.Connection,
                          stock_id: str, date_str: str) -> bool:
    """
    判斷同一標的在最近 REPEAT_SIGNAL_DAYS 個交易日內是否有放量訊號，
    且今日量能不符合 CONTINUE 延續門檻（即非正常量能延續）。

    執行時機：在 detect_continuation 之前，故不依賴今日 CONTINUE DB 資料。
    改為直接判斷今日量是否達 CONTINUE_NORMAL 門檻（>= 1.0 × MA20）：
      - 若達門檻 → 屬量能延續，不標重複
      - 若未達門檻 → 量縮後重現，才標重複
    """
    # 取最近 N 個實際交易日
    recent = con.execute("""
        SELECT date FROM daily_volume
        WHERE date < ?
        ORDER BY date DESC
        LIMIT ?
    """, (date_str, REPEAT_SIGNAL_DAYS)).fetchall()

    if not recent:
        return False

    oldest_trade_date = recent[-1][0]

    # 最近 N 個交易日內是否有放量訊號
    prev_exists = con.execute("""
        SELECT COUNT(*) FROM signals
        WHERE stock_id = ?
          AND date >= ?
          AND date < ?
          AND signal_type NOT IN ('DISPOSED')
    """, (stock_id, oldest_trade_date, date_str)).fetchone()[0]

    if not prev_exists:
        return False  # 近期無訊號，不是重複

    # 今日量是否仍達 CONTINUE 門檻（>= normal_low × MA20）
    # 若達門檻 → 屬量能延續，不標重複
    today_row = con.execute("""
        SELECT CAST(d.volume AS REAL) / m.ma20
        FROM daily_volume d
        JOIN volume_ma m ON d.stock_id = m.stock_id AND d.date = m.date
        WHERE d.stock_id = ? AND d.date = ? AND m.ma20 > 0
    """, (stock_id, date_str)).fetchone()

    if today_row and today_row[0] >= CONTINUE_CFG["normal_low"]:
        return False  # 今日量仍達延續門檻，屬正常延續，不標重複

    return True  # 量縮後重現放量，才是真正意義的重複訊號


def backfill_market_stats(con: sqlite3.Connection) -> None:
    """補算歷史所有交易日的 market_stats（一次性工具）"""
    existing = {r[0] for r in con.execute("SELECT date FROM market_stats").fetchall()}
    all_dates = [r[0] for r in con.execute(
        "SELECT DISTINCT date FROM daily_volume ORDER BY date"
    ).fetchall()]
    todo = [d for d in all_dates if d not in existing]
    log.info(f"補算 market_stats：共 {len(todo)} 個交易日")
    for i, d in enumerate(todo, 1):
        compute_market_stats(con, d)
        if i % 30 == 0:
            log.info(f"  進度：{i}/{len(todo)}")
    log.info("✅ market_stats 補算完成")

# ══════════════════════════════════════════════════════════════════════════════
# 方向一：勝率回測模組
# ══════════════════════════════════════════════════════════════════════════════
def compute_winrate_stats(con: sqlite3.Connection) -> pd.DataFrame:
    """
    從 signals + signal_tracking_days 計算各維度勝率統計。
    D+5  = Stage A day_num=5
    D+10 = Stage B day_num=5
    D+15 = Stage B day_num=10
    """
    # ── Step 1：取得訊號與 D+N 收盤 ─────────────────────────────────────────
    df = pd.read_sql("""
        SELECT
            s.stock_id,
            s.date           AS signal_date,
            s.signal_star,
            s.signal_type,
            s.vol_price_rel,
            s.market,
            s.close_price    AS signal_close,
            d5.close_price   AS close_d5,
            d10.close_price  AS close_d10,
            d15.close_price  AS close_d15
        FROM signals s
        LEFT JOIN signal_tracking_days d5
            ON  s.stock_id  = d5.stock_id
            AND s.date       = d5.signal_date
            AND d5.stage     = 'A'
            AND d5.day_num   = 5
        LEFT JOIN signal_tracking_days d10
            ON  s.stock_id  = d10.stock_id
            AND s.date       = d10.signal_date
            AND d10.stage    = 'B'
            AND d10.day_num  = 5
        LEFT JOIN signal_tracking_days d15
            ON  s.stock_id  = d15.stock_id
            AND s.date       = d15.signal_date
            AND d15.stage    = 'B'
            AND d15.day_num  = 10
        WHERE s.signal_type NOT IN ('DISPOSED')
          AND s.signal_star  IS NOT NULL
          AND s.signal_star  != ''
          AND s.close_price  IS NOT NULL
          AND s.close_price  > 0
    """, con)

    if df.empty:
        log.warning("compute_winrate_stats：signals 表無有效資料")
        return pd.DataFrame()

    # ── Step 2：計算報酬率（以訊號日收盤為基準）───────────────────────────────
    for dn in ["d5", "d10", "d15"]:
        col = f"close_{dn}"
        if col not in df.columns:
            df[col] = np.nan
        # 確保數值欄位為 float（避免 object dtype 導致比較失效）
        df[col]             = pd.to_numeric(df[col],            errors="coerce")
        df["signal_close"]  = pd.to_numeric(df["signal_close"], errors="coerce")

        valid = df[col].notna() & df["signal_close"].notna() & (df["signal_close"] > 0)
        df[f"ret_{dn}"] = np.nan   # 用 np.nan 而非 None，確保 dropna() 能正確移除
        df.loc[valid, f"ret_{dn}"] = (
            (df.loc[valid, col] - df.loc[valid, "signal_close"])
            / df.loc[valid, "signal_close"] * 100
        )
        df[f"win_{dn}"] = df[f"ret_{dn}"] > 0   # NaN > 0 = False，不影響有效值

    # ── 偵錯日誌 ─────────────────────────────────────────────────────────────
    d5_valid = df["ret_d5"].notna().sum()
    log.info(f"勝率回測：總訊號={len(df)}, D+5有效={d5_valid}, "
             f"D+10有效={df['ret_d10'].notna().sum()}, "
             f"D+15有效={df['ret_d15'].notna().sum()}")
    log.info(f"  signal_star 唯一值: {sorted(df['signal_star'].dropna().unique())}")
    log.info(f"  market 唯一值: {sorted(df['market'].dropna().unique())}")

    MIN_SAMPLE = 5

    def _stats(sub: pd.DataFrame) -> dict:
        row: dict = {"樣本數": len(sub)}
        for dn, label in [("d5", "D+5"), ("d10", "D+10"), ("d15", "D+15")]:
            try:
                v = sub[f"ret_{dn}"].dropna()
                n = len(v)
                row[f"{label}_樣本"] = n
                if n >= MIN_SAMPLE:
                    wins = v.apply(lambda x: x > 0).sum()
                    row[f"{label}_勝率%"]    = round(wins / n * 100, 1)
                    row[f"{label}_均報酬%"]  = round(float(v.mean()), 2)
                    row[f"{label}_最大%"]    = round(float(v.max()), 2)
                    row[f"{label}_最小%"]    = round(float(v.min()), 2)
                else:
                    row[f"{label}_勝率%"]    = None
                    row[f"{label}_均報酬%"]  = None
                    row[f"{label}_最大%"]    = None
                    row[f"{label}_最小%"]    = None
            except Exception as e:
                log.warning(f"_stats {label} 計算失敗：{e}")
                row[f"{label}_樣本"]   = None
                row[f"{label}_勝率%"]  = None
                row[f"{label}_均報酬%"] = None
                row[f"{label}_最大%"]  = None
                row[f"{label}_最小%"]  = None
        return row

    results = []

    # ── 分類1：星等（用實際存在的值，避免 hardcode 漏接）───────────────────────
    star_order = ["★★★", "★★", "★"]
    actual_stars = set(df["signal_star"].dropna().unique())
    for star in [s for s in star_order if s in actual_stars]:
        sub = df[df["signal_star"] == star].copy()
        if not sub.empty:
            r = _stats(sub); r["分類"] = "星等"; r["分組"] = star
            results.append(r)

    # ── 分類2：量價關係 ──────────────────────────────────────────────────────
    vpr_order = ["量增價漲 ↑↑", "量增價跌 ↑↓", "量縮價漲 ↓↑", "量縮價跌 ↓↓"]
    actual_vprs = set(df["vol_price_rel"].dropna().unique())
    for vpr in [v for v in vpr_order if v in actual_vprs]:
        sub = df[df["vol_price_rel"] == vpr].copy()
        if not sub.empty:
            r = _stats(sub); r["分類"] = "量價關係"; r["分組"] = vpr
            results.append(r)

    # ── 分類3：訊號類型 ──────────────────────────────────────────────────────
    type_order = ["NORMAL", "RELEASE_STRONG", "RELEASE_NORMAL",
                  "RELEASE_WEAK_SURGE", "RELEASE_HOLD_MA",
                  "CONTINUE_STRONG", "CONTINUE_NORMAL", "CONTINUE_SURGE"]
    actual_types = set(df["signal_type"].dropna().unique())
    for stype in [t for t in type_order if t in actual_types]:
        sub = df[df["signal_type"] == stype].copy()
        if not sub.empty:
            r = _stats(sub); r["分類"] = "訊號類型"; r["分組"] = stype
            results.append(r)

    # ── 分類4：星等 × 量價（交叉）────────────────────────────────────────────
    for star in [s for s in star_order if s in actual_stars]:
        for vpr in [v for v in ["量增價漲 ↑↑", "量增價跌 ↑↓"] if v in actual_vprs]:
            sub = df[(df["signal_star"] == star) & (df["vol_price_rel"] == vpr)].copy()
            if len(sub) >= MIN_SAMPLE:
                r = _stats(sub); r["分類"] = "星等×量價"; r["分組"] = f"{star}　{vpr}"
                results.append(r)

    # ── 分類5：市場別 ────────────────────────────────────────────────────────
    actual_mkts = set(df["market"].dropna().unique())
    for mkt in [m for m in ["TWSE", "TPEX"] if m in actual_mkts]:
        sub = df[df["market"] == mkt].copy()
        if not sub.empty:
            r = _stats(sub); r["分類"] = "市場"; r["分組"] = mkt
            results.append(r)

    if not results:
        log.warning("compute_winrate_stats：無法產生任何分組統計")
        return pd.DataFrame()

    col_order = [
        "分類", "分組", "樣本數",
        "D+5_樣本",  "D+5_勝率%",  "D+5_均報酬%",  "D+5_最大%",  "D+5_最小%",
        "D+10_樣本", "D+10_勝率%", "D+10_均報酬%", "D+10_最大%", "D+10_最小%",
        "D+15_樣本", "D+15_勝率%", "D+15_均報酬%", "D+15_最大%", "D+15_最小%",
    ]
    result_df = pd.DataFrame(results)
    for c in col_order:
        if c not in result_df.columns:
            result_df[c] = None
    return result_df[col_order].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# Excel 輸出（6 個工作表）
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(
    con: sqlite3.Connection,
    signals: pd.DataFrame,
    release_df: pd.DataFrame,
    continuation_df: pd.DataFrame,
    disposed_df: pd.DataFrame,
    target_date: str,
    tracking_data: dict = None,   # {"A": df, "B": df, "C": df, "D": df}
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("未安裝 openpyxl，跳過 Excel 輸出（pip install openpyxl）")
        return None

    FONT_NAME = "Arial"

    def _f(bold=False, color="000000", size=10):
        return Font(name=FONT_NAME, bold=bold, color=color, size=size)

    def _fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _clean(val):
        """None / NaN → 空字串，解決欄位顯示 None 的問題"""
        if val is None:
            return ""
        try:
            if isinstance(val, float) and pd.isna(val):
                return ""
        except Exception:
            pass
        return val

    C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    R  = Alignment(horizontal="right",  vertical="center")

    # 顏色定義
    COL = {
        "header":          "1F3864",
        "star3":           "C0392B",
        "star2":           "E67E22",
        "star1":           "F1C40F",
        "disposed":        "922B21",
        "release_strong":  "B7950B",
        "release_normal":  "7D3C98",
        "release_weak":    "CB4335",
        "release_hold":    "1A7A4A",
        "cont_strong":     "1A5276",
        "cont_normal":     "1E8449",
        "cont_surge":      "CB4335",
        "alt":             "F2F3F4",
        "white":           "FFFFFF",
        # 漲跌幅三階段（紅=漲，綠=跌）
        "up3":   "FADBD8",   # 漲 >3%
        "up5":   "F1948A",   # 漲 >5%
        "up8":   "C0392B",   # 漲 >8%
        "flat":  "F2F3F4",   # 平盤
        "dn3":   "D5F5E3",   # 跌 >3%
        "dn5":   "58D68D",   # 跌 >5%
        "dn8":   "1E8449",   # 跌 >8%
        # 量價關係
        "vpr_up":   "FADBD8",  # 價漲（淡紅）
        "vpr_flat": "F2F3F4",  # 平盤（灰）
        "vpr_dn":   "D5F5E3",  # 價跌（淡綠）
        # 追蹤工作表
        "track_a":  "1A5276",
        "track_b":  "6C3483",
        "track_c":  "784212",
        "track_d":  "1B4332",
    }

    # 取得股票名稱對照表
    name_map = get_stock_name_map(con)

    def _chg_fill(chg_val):
        """根據漲跌幅返回對應填色"""
        try:
            v = float(chg_val)
        except Exception:
            return _fill(COL["white"])
        if v == 0:    return _fill(COL["flat"])
        if v >  8:    return _fill(COL["up8"])
        if v >  5:    return _fill(COL["up5"])
        if v >  3:    return _fill(COL["up3"])
        if v < -8:    return _fill(COL["dn8"])
        if v < -5:    return _fill(COL["dn5"])
        if v < -3:    return _fill(COL["dn3"])
        return _fill(COL["white"])

    def _chg_font(chg_val):
        try:
            v = float(chg_val)
        except Exception:
            return _f()
        if v > 5 or v < -5:
            return _f(bold=True, color="FFFFFF")
        return _f()

    def _vpr_fill(vpr_str):
        if not vpr_str or vpr_str == "平盤":
            return _fill(COL["vpr_flat"])
        if "價漲" in vpr_str:
            return _fill(COL["vpr_up"])
        if "價跌" in vpr_str:
            return _fill(COL["vpr_dn"])
        return _fill(COL["white"])

    def _row_fill(signal_type, signal_star, is_disposed):
        if is_disposed:           return _fill(COL["disposed"])
        if signal_type == "RELEASE_STRONG":     return _fill(COL["release_strong"])
        if signal_type == "RELEASE_NORMAL":     return _fill(COL["release_normal"])
        if signal_type == "RELEASE_WEAK_SURGE": return _fill(COL["release_weak"])
        if signal_type == "RELEASE_HOLD_MA":    return _fill(COL["release_hold"])
        if signal_star == "★★★":  return _fill(COL["star3"])
        if signal_star == "★★":   return _fill(COL["star2"])
        return _fill(COL["star1"])

    def _write_header(ws, title_text, headers, widths, row_num=2):
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = title_text
        ws["A1"].font      = _f(bold=True, color="FFFFFF", size=12)
        ws["A1"].fill      = _fill(COL["header"])
        ws["A1"].alignment = C
        ws.row_dimensions[1].height = 28
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row_num, column=ci, value=h)
            cell.font      = _f(bold=True, color="FFFFFF")
            cell.fill      = _fill(COL["header"])
            cell.alignment = C
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row_num].height = 22

    def _v(val, fmt=None):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "-"
        if fmt == "张":
            return max(0, int(float(val)) // 1000)
        if fmt == "float1":
            return round(float(val), 1)
        if fmt == "float2":
            return round(float(val), 2)
        if fmt == "pct":
            return f"{float(val):+.2f}%"
        return val

    wb = Workbook()

    # ── Sheet 1：今日放量訊號 ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "今日放量訊號"
    ws1.freeze_panes = "A3"

    h1 = ["代號","股票名稱","市場","訊號類型","訊號強度",
          "當日量(張)","前日量(張)","MA5(張)","MA10(張)","MA20(張)","MA60(張)",
          "收盤價","漲跌幅%","量價關係",
          "較前日倍","較MA5倍","較MA10倍","較MA20倍","較MA60倍",
          "大盤漲跌%","情境標記",
          "訊號說明","處置狀態"]
    w1 = [9,12,7,14,10, 12,12,12,12,12,12, 9,9,14, 9,9,9,9,9, 10,12, 48,10]

    total = len(signals)
    _write_header(ws1,
        f"台股成交量異常放大訊號  ◆  統計日期：{target_date}  ◆  共 {total} 檔",
        h1, w1)
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(h1))}2"

    for ri, (_, row) in enumerate(signals.iterrows(), 3):
        stype   = row.get("signal_type", "NORMAL")
        star    = row.get("signal_star",  "")
        is_disp = int(row.get("is_disposed", 0))
        rfill   = _row_fill(stype, star, is_disp)
        alt     = (ri % 2 == 0)

        vals = [
            _clean(row["stock_id"]),
            _clean(name_map.get(row["stock_id"], "")),
            _clean(row.get("market","")),
            SIGNAL_TYPES.get(stype, ("",""))[1],
            _clean(star),
            _v(row.get("volume"),    "张"),
            _v(row.get("vol_prev"),  "张"),
            _v(row.get("ma5"),       "张"),
            _v(row.get("ma10"),      "张"),
            _v(row.get("ma20"),      "张"),
            _v(row.get("ma60"),      "张"),
            _v(row.get("close_price"), "float2"),
            _v(row.get("change_pct"),  "pct"),
            _clean(row.get("vol_price_rel","")),
            _v(row.get("ratio_prev"),  "float1"),
            _v(row.get("ratio_ma5"),   "float1"),
            _v(row.get("ratio_ma10"),  "float1"),
            _v(row.get("ratio_ma20"),  "float1"),
            _v(row.get("ratio_ma60"),  "float1"),
            _v(row.get("market_chg_pct"), "pct"),
            _clean(row.get("context", "")),
            _clean(row.get("signal_detail","")),
            "⚠️ 處置中" if is_disp else "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = L if ci in (22,) else C
            if ci == 4:
                cell.fill = rfill
                cell.font = _f(bold=True, color="FFFFFF")
            elif ci == 5:
                cell.fill = rfill
                cell.font = _f(bold=True, color="FFFFFF")
            elif ci == 13:   # 漲跌幅%
                cell.fill = _chg_fill(row.get("change_pct"))
                cell.font = _chg_font(row.get("change_pct"))
            elif ci == 14:   # 量價關係
                cell.fill = _vpr_fill(val)
                cell.font = _f()
            elif ci == 20:   # 大盤漲跌%
                cell.fill = _chg_fill(row.get("market_chg_pct"))
                cell.font = _chg_font(row.get("market_chg_pct"))
            elif ci == 21:   # 情境標記
                if row.get("context", ""):
                    cell.fill = _fill("FFF3CD")
                    cell.font = _f(bold=True, color="856404")
                else:
                    cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                    cell.font = _f()
            elif ci == 23 and is_disp:
                cell.fill = _fill(COL["disposed"])
                cell.font = _f(bold=True, color="FFFFFF")
            else:
                cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                cell.font = _f()

    # ── Sheet 2：量能延續追蹤 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("量能延續追蹤")
    ws2.freeze_panes = "A3"

    h2 = ["代號","股票名稱","市場","延續類型","放量日期","放量量(張)","放量收盤",
          "次日日期","次日量(張)","次日量/放量","次日收盤","次日漲跌幅%","累計漲幅%",
          "原始訊號類型","原始星等"]
    w2 = [9,12,7,14,12,12,10, 12,12,11,10,11,11, 14,10]

    _write_header(ws2,
        f"量能延續追蹤  ◆  {target_date}  ◆  共 {len(continuation_df)} 檔",
        h2, w2)
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(h2))}2"

    TYPE_FILL_CONT = {
        "CONTINUE_STRONG": COL["cont_strong"],
        "CONTINUE_NORMAL": COL["cont_normal"],
        "CONTINUE_SURGE":  COL["cont_surge"],
    }

    for ri, (_, row) in enumerate(continuation_df.iterrows(), 3):
        ctype = row.get("continue_type","")
        cfill = _fill(TYPE_FILL_CONT.get(ctype, COL["white"]))
        alt   = (ri % 2 == 0)
        vals  = [
            _clean(row["stock_id"]),
            _clean(name_map.get(row["stock_id"], "")),
            _clean(row.get("market","")),
            SIGNAL_TYPES.get(ctype,("",""))[1],
            _clean(row.get("signal_date","")),
            _v(row.get("signal_volume"),       "张"),
            _v(row.get("signal_close"),        "float2"),
            _clean(row.get("continue_date","")),
            _v(row.get("continue_volume"),     "张"),
            _v(row.get("vol_ratio"),           "float2"),
            _v(row.get("continue_close"),      "float2"),
            _v(row.get("continue_change_pct"), "pct"),
            _v(row.get("cumulative_change"),   "pct"),
            SIGNAL_TYPES.get(row.get("signal_type",""),("",""))[1],
            _clean(row.get("signal_star","")),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = C
            if ci == 4:
                cell.fill = cfill
                cell.font = _f(bold=True, color="FFFFFF")
            else:
                cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                cell.font = _f()

    # ── Sheet 3：出關股追蹤 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("出關股追蹤")
    ws3.freeze_panes = "A3"

    h3 = ["代號","股票名稱","市場","出關類型","處置天數",
          "處置前日期","處置前量(張)","處置前收盤",
          "出關量(張)","出關量/處置前","出關收盤","出關漲跌幅%","較處置前價格變化%",
          "訊號強度","訊號說明",
          "期間漲跌幅%","最大單日漲%","最大單日跌%","強漲日數(≥6%)","漲停次數","跌停次數",
          "特殊標記","備註"]
    w3 = [9,12,7,14,10, 12,12,10, 12,12,10,11,12, 10,40, 11,11,11,11,9,9, 20,28]

    _write_header(ws3,
        f"出關股追蹤  ◆  {target_date}  ◆  共 {len(release_df)} 檔",
        h3, w3)
    ws3.auto_filter.ref = f"A2:{get_column_letter(len(h3))}2"

    TYPE_FILL_REL = {
        "RELEASE_STRONG":     COL["release_strong"],
        "RELEASE_NORMAL":     COL["release_normal"],
        "RELEASE_WEAK_SURGE": COL["release_weak"],
        "RELEASE_HOLD_MA":    COL["release_hold"],
    }

    for ri, (_, row) in enumerate(release_df.iterrows(), 3):
        rtype = row.get("release_type","")
        rfill = _fill(TYPE_FILL_REL.get(rtype, COL["white"]))
        alt   = (ri % 2 == 0)
        vals  = [
            _clean(row["stock_id"]),
            _clean(name_map.get(row["stock_id"], "")),
            _clean(row.get("market","")),
            SIGNAL_TYPES.get(rtype,("",""))[1],
            _v(row.get("dispose_days")),
            _clean(row.get("pre_disp_date","")),
            _v(row.get("pre_disp_volume"),     "张"),
            _v(row.get("pre_disp_close"),      "float2"),
            _v(row.get("release_volume"),      "张"),
            _v(row.get("vol_ratio_vs_pre"),    "float2"),
            _v(row.get("release_close"),       "float2"),
            _v(row.get("release_change_pct"),  "pct"),
            _v(row.get("price_change_vs_pre"), "pct"),
            _clean(row.get("signal_star","")),
            _clean(row.get("signal_detail","")),
            # 處置期間分析
            _v(row.get("disposal_period_chg"),  "pct"),
            _v(row.get("disposal_max_gain"),    "pct"),
            _v(row.get("disposal_max_loss"),    "pct"),
            _clean(row.get("disposal_surge_count", "")),
            _clean(row.get("disposal_limit_up",   "")),
            _clean(row.get("disposal_limit_down", "")),
            _clean(row.get("disposal_flag", "")),
            _clean(row.get("extra_note", "")),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = L if ci in (15, 22, 23) else C
            if ci == 4:
                cell.fill = rfill
                cell.font = _f(bold=True, color="FFFFFF")
            elif ci == 23 and row.get("extra_note"):
                # 備註欄有內容時特別標色
                cell.fill = _fill("FFF3CD")
                cell.font = _f(bold=True, color="856404")
            else:
                cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                cell.font = _f()

    # ── Sheet 4：處置股清單 ──────────────────────────────────────────────
    ws4 = wb.create_sheet("處置股清單")
    ws4.freeze_panes = "A2"

    h4 = ["代號","股票名稱","市場","處置起日","處置迄日","處置原因"]
    w4 = [10,12,8,14,14,55]

    for ci, (h, w) in enumerate(zip(h4, w4), 1):
        cell = ws4.cell(row=1, column=ci, value=h)
        cell.font      = _f(bold=True, color="FFFFFF")
        cell.fill      = _fill(COL["header"])
        cell.alignment = C
        cell.border    = _border()
        ws4.column_dimensions[get_column_letter(ci)].width = w

    if not disposed_df.empty:
        for ri, (_, row) in enumerate(disposed_df.iterrows(), 2):
            vals = [
                _clean(row.get("stock_id","")),
                _clean(name_map.get(row.get("stock_id",""), "")),
                _clean(row.get("market","")),
                _clean(row.get("dispose_start","")),
                _clean(row.get("dispose_end","")),
                _clean(row.get("reason","")),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.fill      = _fill(COL["alt"]) if ri%2==0 else _fill(COL["white"])
                cell.font      = _f()
                cell.alignment = L if ci==6 else C
                cell.border    = _border()
    else:
        ws4["A2"] = "目前無處置股資料（或抓取失敗）"
        ws4["A2"].font = _f(color="888888")

    # ── Sheet 5：訊號追蹤工作表（原近期訊號總覽已整合入5日追蹤）──────────────
    STAGE_META = [
        ("A", "5日追蹤",    COL["track_a"], 5),
        ("B", "10日延伸",   COL["track_b"], 10),
        ("C", "15日延伸",   COL["track_c"], 15),
        ("D", "15日再延伸", COL["track_d"], 15),
    ]

    # 各 stage 的顯示保留期：結案/晉升後保留3個交易日（約5個曆法天）
    RETENTION_DAYS = {"A": 12, "B": 22, "C": 32, "D": 32}
    # 取得下一段 stage 對照表
    NEXT_STAGE = {"A": "B", "B": "C", "C": "D"}

    for stage, sheet_label, hdr_color, max_days in STAGE_META:
        ws_t = wb.create_sheet(sheet_label)
        ws_t.freeze_panes = "A3"

        # 保留期截止日
        retention_cutoff = (
            datetime.strptime(target_date, "%Y-%m-%d")
            - timedelta(days=RETENTION_DAYS[stage])
        ).strftime("%Y-%m-%d")

        # 查詢：追蹤中 + 保留期內的已晉升/結案
        t_rows = con.execute("""
            SELECT t.stock_id, t.signal_date, t.signal_close, t.signal_star,
                   t.market, t.stage_start_date, t.stage_start_close, t.status
            FROM signal_tracking t
            WHERE t.stage=?
              AND (
                  t.status = 'tracking'
                  OR t.stage_start_date >= ?
              )
            ORDER BY
                CASE t.status WHEN 'tracking' THEN 0 WHEN 'promoted' THEN 1 ELSE 2 END,
                t.signal_date DESC, t.stock_id
        """, (stage, retention_cutoff)).fetchall()

        # 建立表頭
        day_hdrs = []
        for d in range(1, max_days + 1):
            day_hdrs += [f"D+{d}日期", f"D+{d}收盤", f"D+{d}漲跌%"]

        base_hdrs = ["代號","股票名稱","市場","星等","訊號日","訊號收盤"]
        if stage != "A":
            base_hdrs += ["段起始日","段起始收盤"]
        base_hdrs += ["距訊號累計%"]
        if stage != "A":
            base_hdrs += ["距段起點累計%"]
        base_hdrs += ["最高漲幅%", "距高回落%", "狀態"]

        all_hdrs = base_hdrs + day_hdrs
        base_w = [9, 12, 7, 8, 12, 10]
        if stage != "A":
            base_w += [12, 10]
        base_w += [12]
        if stage != "A":
            base_w += [12]
        base_w += [11, 11, 10]
        day_w = [11, 9, 9] * max_days
        all_w = base_w + day_w

        # 寫標題列
        ws_t.merge_cells(f"A1:{get_column_letter(len(all_hdrs))}1")
        ws_t["A1"] = f"訊號{sheet_label}  ◆  {target_date}  ◆  共 {len(t_rows)} 檔"
        ws_t["A1"].font      = _f(bold=True, color="FFFFFF", size=12)
        ws_t["A1"].fill      = _fill(hdr_color)
        ws_t["A1"].alignment = C
        ws_t.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(zip(all_hdrs, all_w), 1):
            cell = ws_t.cell(row=2, column=ci, value=h)
            cell.font      = _f(bold=True, color="FFFFFF")
            cell.fill      = _fill(hdr_color)
            cell.alignment = C
            cell.border    = _border()
            ws_t.column_dimensions[get_column_letter(ci)].width = w

        ws_t.auto_filter.ref = f"A2:{get_column_letter(len(all_hdrs))}2"

        # 寫資料列
        for ri, trow in enumerate(t_rows, 3):
            sid, sig_date, sig_close, star, market, stg_start, stg_close, status = trow
            alt = (ri % 2 == 0)

            # 取本 stage 的每日資料
            day_data = {r[0]: r for r in con.execute("""
                SELECT day_num, trade_date, close_price, change_pct
                FROM signal_tracking_days
                WHERE stock_id=? AND signal_date=? AND stage=?
                ORDER BY day_num
            """, (sid, sig_date, stage)).fetchall()}

            last_close = None
            for d in range(max_days, 0, -1):
                if d in day_data and day_data[d][2] is not None:
                    last_close = day_data[d][2]
                    break

            # ── 方向三：高點分析 ──────────────────────────────────────────────
            peak_close = None
            for d in range(1, max_days + 1):
                if d in day_data and day_data[d][2] is not None:
                    if peak_close is None or day_data[d][2] > peak_close:
                        peak_close = day_data[d][2]

            # 最高漲幅% = (高點收盤 - 訊號日收盤) / 訊號日收盤
            max_gain = _chg_pct(sig_close, peak_close) if (
                peak_close is not None and sig_close and sig_close > 0
            ) else None

            # 距高回落% = (最新收盤 - 高點收盤) / 高點收盤（負值 = 從高點回落）
            drawdown = _chg_pct(peak_close, last_close) if (
                peak_close is not None and last_close is not None and peak_close > 0
            ) else None

            chg_signal = _chg_pct(sig_close, last_close)
            chg_stage  = _chg_pct(stg_close, last_close) if stage != "A" else None

            # 跨段狀態顯示
            if status == "tracking":
                status_str = "追蹤中"
            elif status == "closed":
                status_str = "結案"
            elif status == "promoted":
                next_s = NEXT_STAGE.get(stage)
                if next_s:
                    next_row = con.execute(
                        "SELECT status FROM signal_tracking WHERE stock_id=? AND signal_date=? AND stage=?",
                        (sid, sig_date, next_s)
                    ).fetchone()
                    if next_row:
                        ns = next_row[0]
                        if ns == "tracking":   status_str = "已晉升(追蹤中)"
                        elif ns == "closed":   status_str = "已晉升(已結案)"
                        elif ns == "promoted": status_str = "已晉升(再晉升)"
                        else:                  status_str = "已晉升"
                    else:
                        status_str = "已晉升"
                else:
                    status_str = "已晉升(最終段)"
            else:
                status_str = status

            base_vals = [
                _clean(sid),
                _clean(name_map.get(sid, "")),
                _clean(market),
                _clean(star),
                _clean(sig_date),
                _v(sig_close, "float2"),
            ]
            if stage != "A":
                base_vals += [_clean(stg_start), _v(stg_close, "float2")]
            base_vals += [
                (f"{chg_signal:+.2f}%" if chg_signal is not None else "-"),
            ]
            if stage != "A":
                base_vals += [(f"{chg_stage:+.2f}%" if chg_stage is not None else "-")]
            base_vals += [
                (f"{max_gain:+.2f}%" if max_gain is not None else "-"),
                (f"{drawdown:+.2f}%" if drawdown is not None else "-"),
                status_str,
            ]

            day_vals = []
            for d in range(1, max_days + 1):
                if d in day_data:
                    _, td, dc, dchg = day_data[d]
                    day_vals += [
                        _clean(td),
                        _v(dc, "float2"),
                        f"{dchg:+.2f}%" if dchg is not None else "-",
                    ]
                else:
                    day_vals += ["", "", ""]

            all_vals = base_vals + day_vals

            for ci, val in enumerate(all_vals, 1):
                cell = ws_t.cell(row=ri, column=ci, value=val)
                cell.border    = _border()
                cell.alignment = C

                # 漲跌幅欄位上色（每3欄一組，第3欄是漲跌%）
                base_col_count = len(base_vals)
                if ci > base_col_count:
                    day_offset = ci - base_col_count - 1
                    if day_offset % 3 == 2:  # 漲跌% 欄
                        try:
                            raw_v = float(str(val).replace("%","").replace("+",""))
                            cell.fill = _chg_fill(raw_v)
                            cell.font = _chg_font(raw_v)
                        except Exception:
                            cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                            cell.font = _f()
                    else:
                        cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                        cell.font = _f()
                elif ci < len(base_vals) and ci >= len(base_vals) - (4 if stage != "A" else 3):
                    # 累計%、最高漲幅%、距高回落% 三欄著色
                    # 距高回落% 在 len(base_vals)-2，值為負代表回落，應反向著色
                    is_drawdown = (ci == len(base_vals) - 2)
                    try:
                        raw_v = float(str(val).replace("%", "").replace("+", ""))
                        if is_drawdown:
                            # 距高回落%：負值（已回落）→ 反向著色，值越小（跌越多）顏色越深綠警示
                            # 正值（創新高）→ 紅色系（少見但屬強勢）
                            cell.fill = _chg_fill(-raw_v)  # 反轉符號使回落顯示為警示色
                            cell.font = _chg_font(-raw_v)
                        else:
                            cell.fill = _chg_fill(raw_v)
                            cell.font = _chg_font(raw_v)
                    except Exception:
                        cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                        cell.font = _f()
                elif ci == len(base_vals):   # 狀態欄
                    if status == "tracking":
                        cell.fill = _fill("D6EAF8")
                        cell.font = _f(bold=True, color="1A5276")
                    elif status == "promoted":
                        if "已結案" in status_str:
                            cell.fill = _fill("FDEBD0")
                            cell.font = _f(bold=True, color="784212")
                        else:
                            cell.fill = _fill("D5F5E3")
                            cell.font = _f(bold=True, color="1E8449")
                    elif status == "closed":
                        cell.fill = _fill("EAECEE")
                        cell.font = _f(color="888888")
                    else:
                        cell.fill = _fill(COL["alt"])
                        cell.font = _f(color="888888")
                else:
                    cell.fill = _fill(COL["alt"]) if alt else _fill(COL["white"])
                    cell.font = _f()


    # ── Sheet 10：勝率回測摘要 ──────────────────────────────────────────────
    wr_df = compute_winrate_stats(con)
    ws_wr = wb.create_sheet("勝率回測")
    ws_wr.freeze_panes = "A2"
    ws_wr.sheet_tab_color = "4472C4"

    if wr_df.empty:
        ws_wr.cell(row=1, column=1, value="資料不足，請執行 --backfill-tracking 補齊歷史追蹤資料")
    else:
        # 欄位標題列
        wr_hdrs = list(wr_df.columns)
        for ci, h in enumerate(wr_hdrs, 1):
            c = ws_wr.cell(row=1, column=ci, value=h)
            c.font      = _f(bold=True, color="FFFFFF")
            c.fill      = _fill("2E4057")
            c.alignment = C
            c.border    = _border()

        # 各分類段的標題底色
        section_colors = {
            "星等":     "D6EAF8",
            "量價關係": "D5F5E3",
            "訊號類型": "FEF9E7",
            "星等×量價":"FADBD8",
            "市場":     "F2F3F4",
        }
        prev_cat = None
        ws_row = 2

        for _, row_series in wr_df.iterrows():
            row_dict = row_series.to_dict()
            cat  = row_dict.get("分類", "")
            bg   = section_colors.get(cat, "FFFFFF")

            # 分類切換時插入分隔標題列
            if cat != prev_cat:
                prev_cat = cat
                ws_wr.insert_rows(ws_row)
                c = ws_wr.cell(row=ws_row, column=1, value=f"── {cat} ──")
                c.font      = _f(bold=True, color="1F3864")
                c.fill      = _fill("BDD7EE")
                c.alignment = L
                for ci2 in range(2, len(wr_hdrs) + 1):
                    c2 = ws_wr.cell(row=ws_row, column=ci2)
                    c2.fill = _fill("BDD7EE")
                ws_row += 1   # 段落標題列佔一列，往下移

            for ci, h in enumerate(wr_hdrs, 1):
                val = row_dict.get(h)
                c   = ws_wr.cell(row=ws_row, column=ci, value=val)
                c.border    = _border()
                c.alignment = C
                if h.endswith("_勝率%") and val is not None:
                    if val >= 60:
                        c.fill = _fill("C8E6C9"); c.font = _f(bold=True, color="1B5E20")
                    elif val >= 50:
                        c.fill = _fill("DCEDC8"); c.font = _f(color="33691E")
                    elif val >= 40:
                        c.fill = _fill("FFF9C4"); c.font = _f(color="F57F17")
                    else:
                        c.fill = _fill("FFCCBC"); c.font = _f(color="BF360C")
                elif h.endswith("_均報酬%") and val is not None:
                    c.fill = _chg_fill(val)
                    c.font = _chg_font(val)
                elif h in ("分類", "分組"):
                    c.fill = _fill(bg)
                    c.font = _f(bold=(h == "分類"))
                else:
                    c.fill = _fill(bg)
                    c.font = _f()

            ws_row += 1   # 資料列佔一列，往下移

        # 欄寬
        col_widths = {"分類": 12, "分組": 20, "樣本數": 8}
        for dn in ["D+5", "D+10", "D+15"]:
            col_widths[f"{dn}_樣本"]   = 7
            col_widths[f"{dn}_勝率%"]  = 9
            col_widths[f"{dn}_均報酬%"]= 10
            col_widths[f"{dn}_最大%"]  = 9
            col_widths[f"{dn}_最小%"]  = 9
        for ci, h in enumerate(wr_hdrs, 1):
            ws_wr.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 10)

    # ── Sheet 10：使用說明 ──────────────────────────────────────────────
    ws6 = wb.create_sheet("使用說明")
    ws6.column_dimensions["A"].width = 95

    guide = [
        ("台股成交量分析系統 v5.3  使用說明",              True,  14, "1F3864", "FFFFFF"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【放量訊號標準（五維度）】",                      True,  11, "EAF2FF", "1F3864"),
        ("  ★★★ 強烈：前日×3.0 / MA5×2.5 / MA10×2.8 / MA20×3.0 / MA60×3.5", False,10,"FFEEEE","CC0000"),
        ("  ★★  中等：前日×2.0 / MA5×1.8 / MA10×1.9 / MA20×2.0 / MA60×2.5", False,10,"FFF3E0","CC6600"),
        ("  ★   注意：前日×1.5 / MA5×1.5 / MA10×1.5 / MA20×1.5 / MA60×2.0", False,10,"FFFDE7","CC9900"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【訊號類型說明（9種）】",                         True,  11, "EAF2FF", "1F3864"),
        ("  一般放量   (NORMAL)           → 正常交易標的出現放量，依星等判斷強度",                   False,10,"FFFFFF","000000"),
        ("  🔒 處置中  (DISPOSED)          → 當日在處置名單，均量可能失真，僅供參考",               False,10,"FFEEEE","922B21"),
        ("  ⚡ 出關強勢(RELEASE_STRONG)    → 出關當日量 ≥ 處置前基準量 × 1.2，資金明確介入",        False,10,"FFF9E7","B7950B"),
        ("  🔓 出關恢復(RELEASE_NORMAL)    → 出關量在基準量 × 0.8~1.2，正常恢復",                  False,10,"F4ECF7","7D3C98"),
        ("  🔥 出關價強(RELEASE_WEAK_SURGE)→ 出關量 < 基準 × 0.8，但漲幅 ≥ 5%，籌碼鎖定",         False,10,"FFEEEE","CB4335"),
        ("  📊 出關守均(RELEASE_HOLD_MA)   → 出關量未達標，但收盤 > MA20，結構未破，值得觀察",      False,10,"E9F7EF","1A7A4A"),
        ("  🚀 強力延續(CONTINUE_STRONG)   → 當日量 ≥ MA20 × 1.5，放量延續中",                    False,10,"E8F4FD","1A5276"),
        ("  ✅ 正常延續(CONTINUE_NORMAL)   → 當日量介於 MA20 × 1.0~1.5，量能維持",               False,10,"EAFAF1","1E8449"),
        ("  🔥 價強量縮(CONTINUE_SURGE)    → 當日量 < MA20，但漲幅 ≥ 5%，籌碼鎖定",               False,10,"FFEEEE","CB4335"),
        ("  ※ CONTINUE 判斷基準：v5.3 起改用今日量 ÷ MA20（原為 ÷ 前日訊號量），接手率大幅提升",   False,10,"FFF9E7","7D3C98"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【情境標記說明（v5.3 新增）】",                   True,  11, "EAF2FF", "1F3864"),
        ("  大盤市況（TWSE全市場個股簡單平均漲跌幅，[lo,hi) 邊界規則）",                          False,10,"F2F2F2","444444"),
        ("  🚀 極漲   ≥ +2.5%　　🌊 強漲  +1.8%~+2.5%　　📈 大漲  +1.0%~+1.8%",               False,10,"FFF3CD","856404"),
        ("  🔼 小漲  +0.3%~+1.0%　↗️ 偏漲  +0.1%~+0.3%　➡️ 中性  -0.1%~+0.1%",              False,10,"F0FFF0","276327"),
        ("  ↘️ 偏跌  -0.3%~-0.1%　🔽 小跌  -1.0%~-0.3%　📉 大跌  -1.8%~-1.0%",              False,10,"FFF0F0","922B21"),
        ("  🌧️ 強跌  -2.5%~-1.8%　🔻 極跌   ≤ -2.5%",                                       False,10,"D6EAF8","1A5276"),
        ("  ⚠️ 重複    → 近 3 個交易日內已有訊號，且今日量 < MA20（量縮後重現），參考價值較低",     False,10,"FFCCBC","BF360C"),
        ("  （以上標記可複合出現，例如：🌊強漲日 ⚠️重複）",                                        False,10,"FFFFFF","888888"),
        ("  大盤漲跌% = TWSE 全市場個股簡單平均漲跌幅（非加權指數），中性 = -0.1%~+0.1%",             False,10,"FFFFFF","888888"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【量價關係說明】",                                True,  11, "EAF2FF", "1F3864"),
        ("  量增價漲 ↑↑ → 最強，主力積極買入",                                                     False,10,"FFFFFF","000000"),
        ("  量增價跌 ↑↓ → 警示，放量下跌可能是出貨訊號",                                           False,10,"FFFFFF","000000"),
        ("  量縮價漲 ↓↑ → 籌碼鎖定，少量買盤推價",                                                 False,10,"FFFFFF","000000"),
        ("  量縮價跌 ↓↓ → 最弱，無力反彈",                                                         False,10,"FFFFFF","000000"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【勝率回測說明（v5.3 新增）】",                   True,  11, "EAF2FF", "1F3864"),
        ("  統計各維度訊號在 D+5 / D+10 / D+15 的歷史勝率與平均報酬",                               False,10,"FFFFFF","000000"),
        ("  D+5  = Stage A 第5個交易日收盤 vs 訊號日收盤",                                         False,10,"FFFFFF","000000"),
        ("  D+10 = Stage B 第5個交易日（= 整體 D+10）",                                           False,10,"FFFFFF","000000"),
        ("  D+15 = Stage B 第10個交易日（= 整體 D+15）",                                          False,10,"FFFFFF","000000"),
        ("  ⚠️ D+10/D+15 樣本只含「晉升 Stage B」的標的（表現穩健的子集），勝率會高於真實整體",    False,10,"FFF9E7","B7950B"),
        ("  勝率著色：≥60% 深綠 / ≥50% 淺綠 / ≥40% 黃 / <40% 橘紅",                             False,10,"FFFFFF","000000"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【欄位單位說明】",                                True,  11, "EAF2FF", "1F3864"),
        ("  成交量欄位單位為「張」（1張 = 1000股）",                                                False,10,"FFFFFF","000000"),
        ("  倍數欄 = 當日量 ÷ 對應基準量，越高代表放量越明顯",                                      False,10,"FFFFFF","000000"),
        ("  出關追蹤：vol_ratio = 出關量 ÷ 處置前基準量",                                          False,10,"FFFFFF","000000"),
        ("  延續追蹤：vol_ratio = 今日量 ÷ 今日MA20（v5.3 修正）",                                False,10,"FFFFFF","000000"),
        ("  距訊號累計% / 距段起點累計% = 相對各基準日的累計報酬率",                                False,10,"FFFFFF","000000"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【CLI 指令說明】",                                True,  11, "EAF2FF", "1F3864"),
        ("  --update               每日更新 + 輸出 Excel（主要指令）",                             False,10,"FFFFFF","000000"),
        ("  --init                 首次初始化，回溯 7 個月歷史資料",                               False,10,"FFFFFF","000000"),
        ("  --backfill YYYYMMDD    補算指定日期的放量訊號",                                        False,10,"FFFFFF","000000"),
        ("  --backfill-tracking    重建所有歷史追蹤資料（更新邏輯後需重跑）",                       False,10,"FFFFFF","000000"),
        ("  --fix-history          補正歷史漲跌幅（除息修正）",                                    False,10,"FFFFFF","000000"),
        ("  --fill-market-stats    補算歷史大盤情境統計",                                          False,10,"FFFFFF","000000"),
        ("  --backfill-history-log 從 DB 重建 history_log.csv（快速摘要）",                       False,10,"FFFFFF","000000"),
        ("  --winrate              輸出獨立的勝率回測 Excel",                                      False,10,"FFFFFF","000000"),
        ("  --stock 2330           查詢個股歷史訊號",                                              False,10,"FFFFFF","000000"),
        ("  --signals              顯示近期訊號列表",                                              False,10,"FFFFFF","000000"),
        ("",                                               False, 10, "FFFFFF", "000000"),
        ("【使用建議】",                                    True,  11, "EAF2FF", "1F3864"),
        ("  ▌ 首次建立（只需一次）",                         True,  10, "F2F3F4", "1F3864"),
        ("  1. python volume_system.py --init",              False, 10, "FFFFFF", "000000"),
        ("     → 回溯 7 個月歷史量價資料（約需 5~10 分鐘）", False, 10, "FFFFFF", "888888"),
        ("  2. python volume_system.py --fill-market-stats", False, 10, "FFFFFF", "000000"),
        ("     → 補算歷史大盤情境統計",                      False, 10, "FFFFFF", "888888"),
        ("  3. python volume_system.py --backfill-tracking", False, 10, "FFFFFF", "000000"),
        ("     → 建立歷史追蹤資料（Stage A~D）",             False, 10, "FFFFFF", "888888"),
        ("  4. python volume_system.py --backfill-history-log",False,10,"FFFFFF","000000"),
        ("     → 產生 history_log.csv（輕量快速摘要）",      False, 10, "FFFFFF", "888888"),
        ("",                                                 False, 10, "FFFFFF", "000000"),
        ("  ▌ 每日例行（收盤後執行）",                        True,  10, "F2F3F4", "1F3864"),
        ("  python volume_system.py --update",               False, 10, "E8F5E9", "000000"),
        ("     → 自動完成：補抓資料、偵測訊號、更新追蹤、輸出 Excel、更新 history_log", False,10,"FFFFFF","888888"),
        ("",                                                 False, 10, "FFFFFF", "000000"),
        ("  ▌ 何時需要重跑 backfill",                         True,  10, "F2F3F4", "1F3864"),
        ("  • 更新程式碼後（訊號偵測邏輯有改動）",            False, 10, "FFFFFF", "000000"),
        ("    → --backfill-tracking 再 --backfill-history-log",False,10,"FFFFFF","888888"),
        ("  • 發現歷史漲跌幅異常（除息日）",                  False, 10, "FFFFFF", "000000"),
        ("    → --fix-history",                              False, 10, "FFFFFF", "888888"),
        ("",                                                 False, 10, "FFFFFF", "000000"),
        ("  ▌ 重點關注邏輯",                                  True,  10, "F2F3F4", "1F3864"),
        ("  最高優先：情境無標記（無強漲日/無重複）＋ ★★★ ＋ 量增價漲↑↑",           False,10,"FFF9E7","B7950B"),
        ("  加分條件：出關強勢（RELEASE_STRONG）或 強力延續（CONTINUE_STRONG）同日出現",False,10,"FFF9E7","B7950B"),
        ("  原因：情境無標記 = 個股獨立動能，不是跟大盤跑；量增價漲 = 主力積極介入",  False,10,"FFFFFF","888888"),
        ("  快速排除：有⚠️重複標記且今日量 < MA20 的標的（量縮重現，動能不足）",      False,10,"FFEEEE","CB4335"),
        ("",                                                 False, 10, "FFFFFF", "000000"),
        ("  ▌ 與 Claude 協作",                               True,  10, "F2F3F4", "1F3864"),
        ("  上傳 history_log.csv 可做跨日趨勢分析（比 Excel 節省 95% 以上 token）",   False,10,"E9F7EF","1A7A4A"),
        ("  需要深入分析個股時，再搭配上傳 tw_volume.db",                             False,10,"FFFFFF","888888"),
        ("  Excel 留給自己本機查看，不需傳給 AI",                                     False,10,"FFFFFF","888888"),
    ]

    for ri, (text, bold, size, bg, fg) in enumerate(guide, 1):
        cell = ws6.cell(row=ri, column=1, value=text)
        cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
        cell.fill      = _fill(bg)
        cell.alignment = L
        ws6.row_dimensions[ri].height = 20

    # 儲存
    fname = OUT_DIR / f"台股放量訊號_{target_date.replace('-','')}.xlsx"
    wb.save(fname)
    log.info(f"Excel 已輸出：{fname}")
    return fname


# ══════════════════════════════════════════════════════════════════════════════
# 主更新流程
# ══════════════════════════════════════════════════════════════════════════════
def manual_update(db_path: str = DB_PATH) -> Path:
    con = init_db(db_path)

    # 1. 補抓缺漏日期
    fetch_history(con, months_back=7)

    # 2. 處置股
    log.info("抓取處置股清單...")
    disp_df = fetch_disposed_stocks()
    if not disp_df.empty:
        today = datetime.today().strftime("%Y-%m-%d")
        existing_disp = {
            r[0] for r in con.execute(
                "SELECT stock_id FROM disposed_stocks WHERE fetch_date=?", (today,)
            ).fetchall()
        }
        new_disp = disp_df[~disp_df["stock_id"].isin(existing_disp)]
        if not new_disp.empty:
            # 逐筆寫入，遇到重複主鍵則忽略
            for _, row in new_disp.iterrows():
                try:
                    con.execute("""
                        INSERT OR IGNORE INTO disposed_stocks
                        (stock_id, fetch_date, dispose_start, dispose_end, reason, market)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        row["stock_id"], row["fetch_date"],
                        row["dispose_start"], row["dispose_end"],
                        row["reason"], row["market"]
                    ))
                except Exception:
                    pass
            con.commit()
        today_str = datetime.today().strftime("%Y-%m-%d")
        active_disp = disp_df[
            (disp_df["dispose_end"] == "") |
            (disp_df["dispose_end"] >= today_str)
            ]
        disposed_ids = set(active_disp["stock_id"].tolist())
        log.info(f"今日有效處置股：{len(disposed_ids)} 檔")
    else:
        disposed_ids = set()

    # 3. 標記歷史處置期間
    mark_disposed_volume(con)

    # 3.5 補正除息/除權造成的漲跌幅誤算（對今日資料）
    latest_date = con.execute("SELECT MAX(date) FROM daily_volume").fetchone()[0]
    if latest_date:
        fix_zero_change_pct(con, latest_date)

    # 4. 計算均量
    compute_moving_averages(con)

    # 5. 取最新交易日
    latest = con.execute("SELECT MAX(date) FROM daily_volume").fetchone()[0]
    if not latest:
        log.warning("資料庫無資料")
        con.close()
        return None

    # 清除今日舊訊號（避免重複）
    con.execute("DELETE FROM signals WHERE date=?",               (latest,))
    con.execute("DELETE FROM release_tracking WHERE release_date=?", (latest,))
    con.execute("DELETE FROM continuation_tracking WHERE continue_date=?", (latest,))
    con.commit()

    # 6. 偵測訊號
    signals     = detect_signals(con, latest, disposed_ids)
    release_df  = detect_release_signals(con, latest, disposed_ids)
    cont_df     = detect_continuation(con, latest)

    # ── 排序：星等 DESC → 成交金額（收盤×量）DESC ────────────────────────────
    # 同時過濾極低流動性股（當日量 < 200張），避免冷門股因倍率誇張佔據前排
    MIN_VOL = 200   # 可調整門檻（張）
    if not signals.empty:
        signals = signals[signals["volume"] >= MIN_VOL].copy()
        signals["turnover"] = signals["close_price"].fillna(0) * signals["volume"].fillna(0)
        star_rank = {"★★★": 3, "★★": 2, "★": 1}
        signals["_sort_star"] = signals["signal_star"].map(star_rank).fillna(0)
        signals = signals.sort_values(
            ["_sort_star", "turnover"],
            ascending=[False, False]
        ).drop(columns=["_sort_star", "turnover"])

    # 7. 更新訊號追蹤
    compute_signal_tracking(con, latest)

    # 8. 輸出 Excel
    out_path = export_excel(con, signals, release_df, cont_df, disp_df, latest)

    # 終端摘要
    print(f"\n{'='*72}")
    print(f"  📊 台股放量訊號報告 v3.0  ◆  {latest}")
    print(f"{'='*72}")
    if not signals.empty:
        print(f"  【放量訊號】 共 {len(signals)} 檔")
        print(f"    ★★★:{(signals['signal_star']=='★★★').sum()} | "
              f"★★:{(signals['signal_star']=='★★').sum()} | "
              f"★:{(signals['signal_star']=='★').sum()} | "
              f"處置中:{(signals['signal_type']=='DISPOSED').sum()}")
    if not release_df.empty:
        print(f"  【出關訊號】 共 {len(release_df)} 檔")
        print(f"    出關強勢:{(release_df['release_type']=='RELEASE_STRONG').sum()} | "
              f"出關恢復:{(release_df['release_type']=='RELEASE_NORMAL').sum()} | "
              f"量縮價強:{(release_df['release_type']=='RELEASE_WEAK_SURGE').sum()} | "
              f"守均:{(release_df['release_type']=='RELEASE_HOLD_MA').sum()}")
    if not cont_df.empty:
        print(f"  【延續訊號】 共 {len(cont_df)} 檔")
        print(f"    強力:{(cont_df['continue_type']=='CONTINUE_STRONG').sum()} | "
              f"正常:{(cont_df['continue_type']=='CONTINUE_NORMAL').sum()} | "
              f"價強:{(cont_df['continue_type']=='CONTINUE_SURGE').sum()}")

    if not signals.empty:
        print(f"\n  前15強訊號：")
        top = signals.head(15)[["stock_id","market","signal_type","signal_star",
                                  "close_price","change_pct","ratio_ma20","vol_price_rel"]]
        top.columns = ["代號","市場","類型","星等","收盤","漲跌%","MA20倍","量價"]
        print(top.to_string(index=False))

    print(f"\n  Excel：{out_path}")
    print(f"{'='*72}\n")

    # ── 每日歷史快照（history_log.csv）────────────────────────────────────────
    _append_history_log(
        target_date  = latest_date,
        signals      = signals,
        continuation = cont_df,
        release      = release_df,
        con          = con,
    )

    con.close()
    return out_path


def _append_history_log(
    target_date:  str,
    signals:      pd.DataFrame,
    continuation: pd.DataFrame,
    release:      pd.DataFrame,
    con:          sqlite3.Connection,
) -> None:
    """
    每次 --update 後，把當日核心統計追加寫入 history_log.csv。
    只記最有診斷價值的欄位，避免冗餘。
    """
    log_path = OUT_DIR / "history_log.csv"

    # ── 訊號統計 ──────────────────────────────────────────────────────────────
    total   = len(signals) if signals is not None and not signals.empty else 0
    s3 = s2 = s1 = 0
    repeat_cnt  = 0
    strong_flag = 0
    if total > 0:
        s3 = (signals["signal_star"] == "★★★").sum()
        s2 = (signals["signal_star"] == "★★").sum()
        s1 = (signals["signal_star"] == "★").sum()
        if "context" in signals.columns:
            repeat_cnt  = signals["context"].str.contains("重複",  na=False).sum()
            strong_flag = signals["context"].str.contains("強漲|極漲", na=False, regex=True).sum()

    # ── 大盤情境 ──────────────────────────────────────────────────────────────
    mkt = get_market_context(con, target_date)
    twse_chg = mkt.get("twse_avg_chg")
    tpex_chg = mkt.get("tpex_avg_chg")
    mkt_flag = mkt.get("market_flag", "")

    # ── 延續偵測 ──────────────────────────────────────────────────────────────
    cont_cnt = len(continuation) if continuation is not None and not continuation.empty else 0

    # ── 出關訊號 ──────────────────────────────────────────────────────────────
    rel_cnt  = len(release)      if release      is not None and not release.empty      else 0

    # ── 組成一列 ──────────────────────────────────────────────────────────────
    row = {
        "日期":         target_date,
        "訊號總數":      total,
        "★★★":         int(s3),
        "★★":          int(s2),
        "★":           int(s1),
        "TWSE均漲跌%":  round(twse_chg, 3) if twse_chg is not None else "",
        "TPEX均漲跌%":  round(tpex_chg, 3) if tpex_chg is not None else "",
        "大盤旗標":      mkt_flag,
        "重複標記數":    int(repeat_cnt),
        "重複率%":      round(repeat_cnt / total * 100, 1) if total > 0 else 0,
        "CONTINUE數":   int(cont_cnt),
        "出關訊號數":    int(rel_cnt),
    }

    new_df = pd.DataFrame([row])

    if log_path.exists():
        try:
            existing = pd.read_csv(log_path, dtype=str, encoding="utf-8-sig")
            if "日期" in existing.columns:
                existing = existing[existing["日期"].astype(str) != str(target_date)]
            else:
                existing = pd.DataFrame()
        except Exception:
            existing = pd.DataFrame()
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df

    combined.to_csv(log_path, index=False, encoding="utf-8-sig")
    log.info(f"歷史快照已更新：{log_path}（共 {len(combined)} 筆）")

def backfill_history_log(con: sqlite3.Connection) -> None:
    """
    從 DB 重建所有歷史交易日的 history_log.csv。
    資料來源：signals、market_stats、continuation_tracking、release_tracking
    """
    log.info("開始補算 history_log...")

    # 取所有有訊號的交易日
    all_dates = [r[0] for r in con.execute(
        "SELECT DISTINCT date FROM signals ORDER BY date"
    ).fetchall()]

    if not all_dates:
        log.warning("signals 表無資料，跳過")
        return

    rows = []
    for date in all_dates:
        # 訊號統計
        total = con.execute(
            "SELECT COUNT(*) FROM signals WHERE date=? AND signal_type NOT IN ('DISPOSED')",
            (date,)
        ).fetchone()[0]
        s3 = con.execute(
            "SELECT COUNT(*) FROM signals WHERE date=? AND signal_star='★★★' AND signal_type NOT IN ('DISPOSED')",
            (date,)
        ).fetchone()[0]
        s2 = con.execute(
            "SELECT COUNT(*) FROM signals WHERE date=? AND signal_star='★★' AND signal_type NOT IN ('DISPOSED')",
            (date,)
        ).fetchone()[0]
        s1 = con.execute(
            "SELECT COUNT(*) FROM signals WHERE date=? AND signal_star='★' AND signal_type NOT IN ('DISPOSED')",
            (date,)
        ).fetchone()[0]

        # 大盤情境（從 market_stats）
        mkt_row = con.execute(
            "SELECT twse_avg_chg, tpex_avg_chg, market_flag FROM market_stats WHERE date=?",
            (date,)
        ).fetchone()
        twse_chg = round(mkt_row[0], 3) if mkt_row and mkt_row[0] is not None else ""
        tpex_chg = round(mkt_row[1], 3) if mkt_row and mkt_row[1] is not None else ""
        mkt_flag = mkt_row[2] if mkt_row and mkt_row[2] else ""

        # CONTINUE 數量
        cont_cnt = con.execute(
            "SELECT COUNT(*) FROM continuation_tracking WHERE continue_date=?",
            (date,)
        ).fetchone()[0]

        # 出關訊號數
        rel_cnt = con.execute(
            "SELECT COUNT(*) FROM release_tracking WHERE release_date=?",
            (date,)
        ).fetchone()[0]

        rows.append({
            "日期":        date,
            "訊號總數":     total,
            "★★★":        s3,
            "★★":         s2,
            "★":          s1,
            "TWSE均漲跌%": twse_chg,
            "TPEX均漲跌%": tpex_chg,
            "大盤旗標":     mkt_flag,
            "重複標記數":   "",      # 歷史無法重建（context 不存 DB）
            "重複率%":      "",
            "CONTINUE數":  cont_cnt,
            "出關訊號數":   rel_cnt,
        })

    df = pd.DataFrame(rows)
    log_path = OUT_DIR / "history_log.csv"
    df.to_csv(log_path, index=False, encoding="utf-8-sig")
    log.info(f"✅ history_log 補算完成：{len(df)} 個交易日 → {log_path}")



# ══════════════════════════════════════════════════════════════════════════════
# 查詢工具
# ══════════════════════════════════════════════════════════════════════════════
def query_stock(stock_id: str, days: int = 60, db_path: str = DB_PATH) -> pd.DataFrame:
    con = sqlite3.connect(db_path)
    df  = pd.read_sql("""
        SELECT v.date,
               v.volume/1000        AS 量_張,
               v.close_price        AS 收盤,
               v.change_pct         AS 漲跌幅,
               v.is_disposed        AS 處置中,
               ROUND(m.vol_prev/1000) AS 前日量_張,
               ROUND(m.ma5/1000)    AS MA5_張,
               ROUND(m.ma10/1000)   AS MA10_張,
               ROUND(m.ma20/1000)   AS MA20_張,
               ROUND(m.ma60/1000)   AS MA60_張
        FROM daily_volume v
        JOIN volume_ma m ON v.stock_id=m.stock_id AND v.date=m.date
        WHERE v.stock_id=?
        ORDER BY v.date DESC LIMIT ?
    """, con, params=(stock_id, days))
    con.close()
    return df


def query_signals(days: int = 30, min_star: int = 1, db_path: str = DB_PATH) -> pd.DataFrame:
    star_map = {1:"★", 2:"★★", 3:"★★★"}
    stars    = [star_map[i] for i in range(min_star, 4)]
    cutoff   = (datetime.today() - timedelta(days=days)).strftime("%Y-%m-%d")
    ph       = ",".join("?"*len(stars))
    con      = sqlite3.connect(db_path)
    df       = pd.read_sql(f"""
        SELECT stock_id, date, market, signal_type, signal_star,
               close_price, change_pct, ratio_ma20, vol_price_rel, signal_detail, is_disposed
        FROM signals
        WHERE date>=? AND signal_star IN ({ph})
        ORDER BY date DESC, signal_star DESC, ratio_ma20 DESC
    """, con, params=[cutoff]+stars)
    con.close()
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 補跑歷史訊號（補 signals 表，不輸出 Excel）
# ══════════════════════════════════════════════════════════════════════════════
def backfill_signals(target_date_str: str, db_path: str = DB_PATH):
    """
    補寫指定日期的放量訊號到 signals 表。
    用途：首次上線後，補寫前一交易日的訊號，讓下次 --update 能產出延續追蹤。

    target_date_str：格式 'YYYY-MM-DD' 或 'YYYYMMDD'
    """
    # 日期格式正規化
    raw = target_date_str.strip().replace("-", "")
    if len(raw) != 8 or not raw.isdigit():
        log.error(f"日期格式錯誤，請用 YYYYMMDD 或 YYYY-MM-DD：{target_date_str}")
        return
    target_date = f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"

    con = init_db(db_path)

    # 確認該日有資料
    row = con.execute(
        "SELECT COUNT(*) FROM daily_volume WHERE date=?", (target_date,)
    ).fetchone()
    if not row or row[0] == 0:
        log.warning(f"{target_date} 在 daily_volume 無資料，先執行 --update 補抓")
        con.close()
        return

    # 確認均量存在
    mrow = con.execute(
        "SELECT COUNT(*) FROM volume_ma WHERE date=?", (target_date,)
    ).fetchone()
    if not mrow or mrow[0] == 0:
        log.info("均量資料不存在，先重算...")
        compute_moving_averages(con)

    # 取該日有效處置股（從 disposed_stocks 中判斷）
    disposed_rows = con.execute("""
        SELECT DISTINCT stock_id FROM disposed_stocks
        WHERE dispose_start <= ? AND (dispose_end = '' OR dispose_end >= ?)
    """, (target_date, target_date)).fetchall()
    disposed_ids = {r[0] for r in disposed_rows}
    log.info(f"{target_date} 對應處置股：{len(disposed_ids)} 檔")

    # 清除該日舊訊號（避免重複）
    deleted = con.execute(
        "DELETE FROM signals WHERE date=?", (target_date,)
    ).rowcount
    if deleted:
        log.info(f"清除 {target_date} 舊訊號 {deleted} 筆")
    con.commit()

    # 偵測並寫入訊號
    sig_df = detect_signals(con, target_date, disposed_ids)

    con.close()

    if sig_df.empty:
        log.warning(f"{target_date} 無放量訊號（資料可能不足）")
    else:
        log.info(f"✅ {target_date} 補寫完成：{len(sig_df)} 筆訊號已存入 DB")
        log.info(f"   請接著執行 --update 即可產出含延續追蹤的完整報告")



if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="台股成交量分析系統 v3.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("--init",     action="store_true", help="首次初始化")
    parser.add_argument("--update",   action="store_true", help="手動更新並輸出Excel")
    parser.add_argument("--backfill", type=str, default=None,
                        help="補寫指定日期訊號到DB，例：--backfill 20260429")
    parser.add_argument("--backfill-tracking", action="store_true",
                        help="從2026-04-06回溯補寫訊號追蹤資料")
    parser.add_argument("--fix-history", action="store_true",
                        help="補正歷史資料中因除息/除權造成的漲跌幅誤算")
    parser.add_argument("--stock",    type=str,  default=None, help="查詢個股，例：--stock 2330")
    parser.add_argument("--days",     type=int,  default=60,   help="查詢天數（預設60）")
    parser.add_argument("--signals",  action="store_true",     help="顯示近期訊號")
    parser.add_argument("--winrate",  action="store_true",     help="輸出勝率回測統計")
    parser.add_argument("--fill-market-stats", action="store_true",
                        help="補算所有歷史交易日的大盤情境統計（market_stats）")
    parser.add_argument("--backfill-history-log", action="store_true",
                        help="從 DB 重建所有歷史交易日的 history_log.csv")
    parser.add_argument("--star",     type=int,  default=1,    help="最低星等 1/2/3")
    args = parser.parse_args()

    if args.init:
        con = init_db()
        fetch_history(con, months_back=7)
        compute_moving_averages(con)
        con.close()
        log.info("✅ 初始化完成！請執行 --update 產出第一份報告。")

    elif args.update:
        manual_update()

    elif args.backfill:
        backfill_signals(args.backfill)

    elif args.backfill_tracking:
        con = init_db()
        backfill_signal_tracking(con, start_date="2026-04-06")
        con.close()
        log.info("✅ 追蹤資料回溯完成！請執行 --update 產出報告。")

    elif args.fix_history:
        con = init_db()
        all_dates = [r[0] for r in con.execute(
            "SELECT DISTINCT date FROM daily_volume ORDER BY date"
        ).fetchall()]
        log.info(f"開始補正歷史漲跌幅（共 {len(all_dates)} 個交易日）...")
        total_fixed = 0
        for i, d in enumerate(all_dates, 1):
            before = con.execute(
                "SELECT COUNT(*) FROM daily_volume WHERE date=? AND change_pct=0", (d,)
            ).fetchone()[0]
            fix_zero_change_pct(con, d)
            after = con.execute(
                "SELECT COUNT(*) FROM daily_volume WHERE date=? AND change_pct=0", (d,)
            ).fetchone()[0]
            fixed = before - after
            total_fixed += fixed
            if fixed > 0 or i % 20 == 0:
                log.info(f"  [{i}/{len(all_dates)}] {d}：補正 {fixed} 筆")
        con.close()
        log.info(f"✅ 歷史補正完成，共修正 {total_fixed} 筆")

    elif args.stock:
        df = query_stock(args.stock, days=args.days)
        print(f"\n── {args.stock} 近 {args.days} 日 ──")
        print(df.to_string(index=False))

    elif args.signals:
        df = query_signals(days=args.days, min_star=args.star)
        print(f"\n── 近 {args.days} 日訊號（≥{args.star}★）共 {len(df)} 筆 ──")
        print(df.to_string(index=False))

    elif args.winrate:
        con = init_db()
        wr = compute_winrate_stats(con)
        con.close()
        if wr.empty:
            print("資料不足，請先執行 --backfill-tracking 補齊歷史追蹤資料")
        else:
            pd.set_option("display.max_columns", None)
            pd.set_option("display.width", 200)
            pd.set_option("display.float_format", lambda x: f"{x:+.2f}%" if isinstance(x, float) else str(x))
            # 只印重點欄位
            show_cols = ["分類","分組","樣本數",
                         "D+5_勝率%","D+5_均報酬%",
                         "D+10_勝率%","D+10_均報酬%",
                         "D+15_勝率%","D+15_均報酬%"]
            print("\n── 勝率回測統計 ──")
            print(wr[show_cols].to_string(index=False))
            # 同時輸出獨立 Excel
            from openpyxl import Workbook as _WB
            out = OUT_DIR / f"勝率回測_{datetime.today().strftime('%Y%m%d')}.xlsx"
            wr.to_excel(out, index=False, sheet_name="勝率回測")
            log.info(f"Excel 已輸出：{out}")

    elif args.fill_market_stats:
        con = init_db()
        backfill_market_stats(con)
        con.close()

    elif args.backfill_history_log:
        con = init_db()
        backfill_history_log(con)
        con.close()

    else:
        parser.print_help()
