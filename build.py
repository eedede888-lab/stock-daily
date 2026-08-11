#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build.py — 把 data/ 下每天/每週的 Excel 轉成靜態網站用的 JSON。"""
import os, re, json, glob, sys
import openpyxl

SKIP_EXISTING = "--force" not in sys.argv
ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(ROOT, "data")
WEEKLY_DIR = os.path.join(DATA_DIR, "weekly")   # data/weekly/YYYYMMDD-YYYYMMDD/
OUT_DIR = os.path.join(ROOT, "site", "data")

def load_name_overrides():
    p = os.path.join(ROOT, "stock_names.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                return {str(k): str(v) for k, v in json.load(f).items()}
        except Exception:
            pass
    return {}

NAME_OVERRIDES = load_name_overrides()


def rows_of(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); return []
    ws = wb[sheet]
    out = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close(); return out


def find_header(rows, key="代號"):
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if key in cells:
            return i, cells
    return None, None


def to_records(rows, key="代號"):
    hi, header = find_header(rows, key)
    if hi is None:
        return []
    recs = []
    for r in rows[hi + 1:]:
        if r is None:
            continue
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        first = str(r[0]).strip() if r[0] is not None else ""
        if first.startswith("──") or first.startswith("【"):
            continue
        rec = {}
        for h, v in zip(header, r):
            if h == "":
                continue
            rec[h] = v
        recs.append(rec)
    return recs


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "").replace("%", "")
    if s == "" or s == "-":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def build_name_map(market_file):
    name_map = {}
    if not market_file:
        return name_map
    small_sheets = ["今日放量訊號", "量能延續追蹤", "出關股追蹤", "處置股清單"]
    wb = openpyxl.load_workbook(market_file, data_only=True, read_only=True)
    for sh in small_sheets:
        if sh not in wb.sheetnames:
            continue
        rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
        hi, header = find_header(rows, "代號")
        if hi is None or "股票名稱" not in header:
            continue
        ci, ni = header.index("代號"), header.index("股票名稱")
        for r in rows[hi + 1:]:
            if r and r[ci] is not None and r[ni]:
                name_map[str(r[ci]).strip()] = str(r[ni]).strip()
    wb.close(); return name_map


# 市場別 → Yahoo 股市代號後綴（上市 TWSE→TW，上櫃 TPEX→TWO）。
def yahoo_suffix(market):
    m = str(market or "").strip().upper()
    if m in ("TPEX", "OTC", "上櫃", "櫃買"):
        return "TWO"
    return "TW"  # 預設視為上市


def build_market_map(market_file):
    """代號 → Yahoo 後綴（TW / TWO），掃描各小分頁的「市場」欄。"""
    mkt_map = {}
    if not market_file:
        return mkt_map
    sheets = ["今日放量訊號", "量能延續追蹤", "出關股追蹤", "處置股清單"]
    wb = openpyxl.load_workbook(market_file, data_only=True, read_only=True)
    for sh in sheets:
        if sh not in wb.sheetnames:
            continue
        rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
        hi, header = find_header(rows, "代號")
        if hi is None or "市場" not in header:
            continue
        ci, mi = header.index("代號"), header.index("市場")
        for r in rows[hi + 1:]:
            if r and r[ci] is not None and r[mi]:
                mkt_map[str(r[ci]).strip()] = yahoo_suffix(r[mi])
    wb.close(); return mkt_map




def market_json_ok(path):
    if not os.path.exists(path):
        return False
    try:
        with open(path, encoding="utf-8") as f:
            json.load(f)
        return True
    except Exception:
        return False

def stock_json_ok(path):
    """既有檔案需能正確解析且含 name 才視為完成，否則重建。"""
    if not os.path.exists(path):
        return False
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        return bool(d.get("name"))
    except Exception:
        return False

def process_market(market_file, out_path):
    """主檔：signals/continuation/release/disposed/winrate（小資料）。"""
    MAIN_SHEETS = {
        "signals":      ("今日放量訊號",   "代號"),
        "continuation": ("量能延續追蹤",   "代號"),
        "release":      ("出關股追蹤",     "代號"),
        "disposed":     ("處置股清單",     "代號"),
        "winrate":      ("勝率回測",       "分類"),
    }
    data = {}
    for key, (sheet, hkey) in MAIN_SHEETS.items():
        recs = to_records(rows_of(market_file, sheet), hkey)
        for rec in recs:
            for k, v in rec.items():
                rec[k] = num(v)
        data[key] = recs
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    return {k: len(v) for k, v in data.items()}


TRACK_SHEETS = {
    "track5":   ("5日追蹤",    "代號"),
    "track10":  ("10日延伸",   "代號"),
    "track15":  ("15日延伸",   "代號"),
    "track15b": ("15日再延伸", "代號"),
}

def process_market_track(market_file, key, out_path):
    """單張追蹤表獨立輸出（每張各自一個檔案，避免合併後超過 25MB）。"""
    sheet, hkey = TRACK_SHEETS[key]
    recs = to_records(rows_of(market_file, sheet), hkey)
    for rec in recs:
        for k, v in rec.items():
            rec[k] = num(v)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(recs, f, ensure_ascii=False, separators=(",", ":"))
    return len(recs)


def process_stock(code, name, files, out_path):
    daily, analysis = files.get("daily"), files.get("analysis")
    out = {"code": code, "name": name, "buy_top": [], "sell_top": [], "price_volume": [], "broker_detail": []}
    if daily:
        out["buy_top"] = [{k: num(v) for k, v in r.items()} for r in to_records(rows_of(daily, "買進前20"), "券商")]
        out["sell_top"] = [{k: num(v) for k, v in r.items()} for r in to_records(rows_of(daily, "賣出前20"), "券商")]
    if analysis:
        out["price_volume"] = [{k: num(v) for k, v in r.items()} for r in to_records(rows_of(analysis, "買賣價量與家數"), "股價")]
        keep = ["股價", "券商", "買進股數", "賣出股數", "買進占比", "賣出占比", "最高買進", "最高賣出"]
        det = []
        for r in to_records(rows_of(analysis, "券商明細"), "股價"):
            det.append({k: num(r.get(k)) for k in keep if k in r})
        out["broker_detail"] = det
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    return {"buy": len(out["buy_top"]), "sell": len(out["sell_top"]),
            "pv": len(out["price_volume"]), "detail": len(out["broker_detail"])}


def _intify(x):
    """整數值存成 int，否則維持 float（比照 num() 的數字呈現）。"""
    if isinstance(x, float) and x.is_integer():
        return int(x)
    return x


def _strip_cn(s):
    """券商正規化：只留中文（去掉分點代碼/全形空白），對應夥伴 日報.py 的 groupby 粒度。"""
    return re.sub(r"[^一-龥]", "", str(s))


def read_bsr_csv(csv_path):
    """讀 bsr_fetch.py 產出的乾淨 CSV → list：券商(raw)/股價(float)/買進股數/賣出股數(int)。"""
    import csv as _csv
    recs = []
    raw = None
    used_enc = None
    for enc in ("utf-8-sig", "cp950", "big5", "gbk"):
        try:
            with open(csv_path, encoding=enc) as f:
                raw = f.read()
            used_enc = enc
            break
        except UnicodeDecodeError:
            continue
    if raw is None:
        print(f"  [警告] 無法判斷編碼，略過此檔: {csv_path}")
        return recs
    if used_enc != "utf-8-sig":
        print(f"  [警告] {csv_path} 非 UTF-8，改用 {used_enc} 讀取")
    from io import StringIO
    with StringIO(raw) as f:
        for r in _csv.DictReader(f):
            try:
                px = round(float(str(r["股價"]).replace(",", "")), 2)
            except (ValueError, KeyError, TypeError):
                continue
            def _i(v):
                v = str(v or "").replace(",", "").strip()
                try: return int(float(v)) if v else 0
                except ValueError: return 0
            recs.append({"券商": str(r.get("券商", "")).strip(), "股價": px,
                         "買進股數": _i(r.get("買進股數")), "賣出股數": _i(r.get("賣出股數"))})
    return recs


def process_stock_from_bsr(code, name, csv_path, out_path):
    """直接從 BSR 分點明細 CSV 算出網站要的 buy_top/sell_top/price_volume/broker_detail。
    公式同夥伴的 日報.py + 量價分析插入png圖彙總.py，但讀「完整」資料（左右雙欄都在，
    Σ買≈Σ賣），修正了舊流程 usecols=[0..4] 只讀左半、漏掉約一半紀錄的問題。"""
    recs = read_bsr_csv(csv_path)
    out = {"code": code, "name": name, "buy_top": [], "sell_top": [],
           "price_volume": [], "broker_detail": []}
    if not recs:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        return {"buy": 0, "sell": 0, "pv": 0, "detail": 0}

    # ── buy_top / sell_top：依「券商總公司」(去非中文) 聚合，均價=Σ金額/Σ股數，量轉張 ──
    agg = {}
    for r in recs:
        b = _strip_cn(r["券商"]) or r["券商"]
        a = agg.setdefault(b, {"bq": 0, "bv": 0.0, "sq": 0, "sv": 0.0})
        a["bq"] += r["買進股數"]; a["bv"] += r["股價"] * r["買進股數"]
        a["sq"] += r["賣出股數"]; a["sv"] += r["股價"] * r["賣出股數"]
    brokers = []
    for b, a in agg.items():
        brokers.append({
            "券商": b,
            "buy_total_qty": _intify(a["bq"] / 1000), "buy_total_value": _intify(a["bv"]),
            "buy_avg_price": (a["bv"] / a["bq"]) if a["bq"] else 0,
            "sell_total_qty": _intify(a["sq"] / 1000), "sell_total_value": _intify(a["sv"]),
            "sell_avg_price": (a["sv"] / a["sq"]) if a["sq"] else 0,
            "buy_sell_diff": _intify((a["bq"] - a["sq"]) / 1000),
        })
    out["buy_top"] = sorted(brokers, key=lambda x: x["buy_total_qty"], reverse=True)[:20]
    out["sell_top"] = sorted(brokers, key=lambda x: x["sell_total_qty"], reverse=True)[:20]

    # ── 各價位總量 + 家數（買賣價量與家數）──
    pv = {}
    for r in recs:
        p = pv.setdefault(r["股價"], {"買進股數": 0, "賣出股數": 0, "買進家數": 0, "賣出家數": 0})
        p["買進股數"] += r["買進股數"]; p["賣出股數"] += r["賣出股數"]
        if r["買進股數"] > 0: p["買進家數"] += 1
        if r["賣出股數"] > 0: p["賣出家數"] += 1

    # ── 各(價位,券商raw)彙總 → 占比 / 最高買賣標記（券商明細）──
    pb = {}
    for r in recs:
        k = (r["股價"], r["券商"])
        v = pb.setdefault(k, [0, 0]); v[0] += r["買進股數"]; v[1] += r["賣出股數"]
    max_buy, max_sell = {}, {}
    for (p, _b), (bq, sq) in pb.items():
        if bq > max_buy.get(p, 0): max_buy[p] = bq
        if sq > max_sell.get(p, 0): max_sell[p] = sq

    out["price_volume"] = [{"股價": p, **pv[p],
                            "最高買進股數": max_buy.get(p, 0), "最高賣出股數": max_sell.get(p, 0)}
                           for p in sorted(pv)]
    det = []
    for (p, b), (bq, sq) in sorted(pb.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        tb, ts = pv[p]["買進股數"], pv[p]["賣出股數"]
        det.append({"股價": p, "券商": b, "買進股數": bq, "賣出股數": sq,
                    "買進占比": round(bq / (tb or 1) * 100, 3),
                    "賣出占比": round(sq / (ts or 1) * 100, 3),
                    "最高買進": bq == max_buy.get(p, 0) and bq > 0,
                    "最高賣出": sq == max_sell.get(p, 0) and sq > 0})
    out["broker_detail"] = det

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    return {"buy": len(out["buy_top"]), "sell": len(out["sell_top"]),
            "pv": len(out["price_volume"]), "detail": len(out["broker_detail"])}


def classify(files_in_dir):
    """日報資料夾用：分類 market / daily / analysis / bsr，忽略 charts。
    優先使用 BSR CSV（完整資料）；同代號若另有日報/分析 xlsx 則作為後備。"""
    market = None
    stocks = {}
    for path in files_in_dir:
        fn = os.path.basename(path)
        if fn.startswith("台股放量訊號"):
            market = path; continue
        m = re.match(r"^(\d{4})", fn)
        if not m:
            continue
        code = m.group(1)
        if "charts" in fn.lower():   # 忽略 charts.xlsx（已無用途）
            continue
        s = stocks.setdefault(code, {})
        if fn.lower().endswith("_bsr.csv"):
            s["bsr"] = path
        elif "日報" in fn:
            s["daily"] = path
        elif "分析結果" in fn:
            s["analysis"] = path
    return market, stocks


def classify_weekly(files_in_dir):
    """週報資料夾用：分類 weekly（週報）/ volume_avg（大量與均價）。"""
    stocks = {}
    for path in files_in_dir:
        fn = os.path.basename(path)
        m = re.match(r"^(\d{4})", fn)
        if not m:
            continue
        code = m.group(1)
        if "charts" in fn.lower():
            continue
        s = stocks.setdefault(code, {})
        if "週報" in fn:
            s["weekly"] = path
        elif "大量與均價" in fn:
            s["volume_avg"] = path
    return stocks


def process_weekly(code, name, files, out_path):
    """讀取週報的買進前20/賣出前20，結構與 process_stock 相同。"""
    weekly = files.get("weekly")
    out = {"code": code, "name": name, "buy_top": [], "sell_top": []}
    if weekly:
        out["buy_top"] = [{k: num(v) for k, v in r.items()}
                          for r in to_records(rows_of(weekly, "買進前20"), "券商")]
        out["sell_top"] = [{k: num(v) for k, v in r.items()}
                          for r in to_records(rows_of(weekly, "賣出前20"), "券商")]
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    return {"buy": len(out["buy_top"]), "sell": len(out["sell_top"])}


def process_volume_avg(code, name, files, out_path):
    """解析「大量與均價」Excel，輸出每日摘要 + 前10明細。

    工作表版面（工作表1）：
      Row 3  : 大欄標題（大量 / 前10買量 / 前10買進均價 / 大量 / 前10賣量…）
      Row 4  : 細欄標題（股價 / 買進股數 / 賣出股數 / 買進家數 / 賣出家數 / 最高買進 / 最高賣出 / 收盤價）
      Row 6~ : 每日摘要（日期, 大量股價, 買進股數, 賣出股數, 買進家數, 賣出家數,
                        最高買進股數, 最高賣出股數, 收盤價,
                        前10買量, 前10買進均價, None,
                        大量(同前), 前10賣量, 前10賣出均價）
      空行後
      Row 14 : 各日期欄位（col 0,4,8,12,16 為 datetime）
      Row 15 : 買/價/賣/價（每日4欄重複）
      Row 16+: 前10券商明細（每列20欄，每5欄一天）
      Row 27 : 各日前10買量/賣量小計
      Row 28 : 各日前10均價
    """
    va_file = files.get("volume_avg")
    out = {"code": code, "name": name, "daily_summary": [], "top10_detail": []}
    if not va_file:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        return {"days": 0}

    wb = openpyxl.load_workbook(va_file, data_only=True, read_only=True)
    if "工作表1" not in wb.sheetnames:   # 版面不符就回空，比照本檔其他解析的容錯做法
        wb.close()
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        return {"days": 0}
    rows = [list(r) for r in wb["工作表1"].iter_rows(values_only=True)]
    wb.close()

    # --- 每日摘要（row index 6~，遇 None 日期停止）---
    summary_cols = ["date", "price", "buy_qty", "sell_qty",
                    "buy_cnt", "sell_cnt", "max_buy_qty", "max_sell_qty", "close",
                    "top10_buy_qty", "top10_buy_avg", None,
                    None, "top10_sell_qty", "top10_sell_avg"]
    daily = []
    for r in rows[6:]:
        if r[0] is None:
            break
        rec = {}
        for i, col in enumerate(summary_cols):
            if col is None or i >= len(r):
                continue
            v = r[i]
            if col == "date":
                rec[col] = v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
            else:
                rec[col] = num(v)
        daily.append(rec)
    out["daily_summary"] = daily

    # --- 找前10明細區塊（日期列，row 14 = index 14）---
    # 每天佔4欄（買量, 買均價, 賣量, 賣均價），5天共20欄
    if len(rows) <= 16:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        return {"days": len(daily)}
    date_row = rows[14]
    dates_order = []
    for i in range(0, 20, 4):
        if i >= len(date_row):
            break
        v = date_row[i]
        if v is not None:
            d = v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
            dates_order.append((d, i))

    detail = {d: [] for d, _ in dates_order}
    for r in rows[16:]:
        if all(c is None for c in r):
            break
        for d, base in dates_order:
            if len(r) <= base + 3:
                continue
            buy_qty, buy_avg, sell_qty, sell_avg = r[base], r[base+1], r[base+2], r[base+3]
            if buy_qty is None and sell_qty is None:
                continue
            detail[d].append({
                "buy_qty":  num(buy_qty),
                "buy_avg":  num(buy_avg),
                "sell_qty": num(sell_qty),
                "sell_avg": num(sell_avg),
            })

    out["top10_detail"] = [{"date": d, "brokers": detail[d]} for d, _ in dates_order]

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    return {"days": len(daily)}


def emit_js_wrappers():
    """把每個 .json 另存成同名 .js，供 file:// 直接以 <script> 載入（免伺服器）。"""
    import glob as _g
    for jp in _g.glob(os.path.join(OUT_DIR, "**", "*.json"), recursive=True):
        key = os.path.relpath(jp, OUT_DIR)[:-5].replace(os.sep, "/")
        with open(jp, encoding="utf-8") as f:
            content = f.read()
        with open(jp[:-5] + ".js", "w", encoding="utf-8") as f:
            f.write("window.__DATAREG&&window.__DATAREG(" + json.dumps(key) + "," + content + ");")


def process_history_log(csv_path, out_path):
    """讀取 history_log.csv（累積快照），輸出為 site/data/history_log.json。
    取最新那份 CSV（放在日期最新的資料夾裡），全量覆蓋輸出。"""
    import csv as _csv
    rows = []
    for enc in ("utf-8-sig", "utf-8", "cp950"):
        try:
            with open(csv_path, encoding=enc, newline="") as f:
                reader = _csv.DictReader(f)
                rows = list(reader)
            break
        except (UnicodeDecodeError, FileNotFoundError):
            continue
    if not rows:
        return 0
    # 數值欄自動轉型
    NUM_COLS = {"訊號總數", "★★★", "★★", "★", "TWSE均漲跌%", "TPEX均漲跌%",
                "重複標記數", "重複率%", "CONTINUE數", "出關訊號數"}
    out = []
    for r in rows:
        rec = {}
        for k, v in r.items():
            if k in NUM_COLS:
                rec[k] = num(v)
            else:
                rec[k] = v
        out.append(rec)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    return len(out)


def build_global_market_map(date_dirs):
    """跨所有日期累積 代號→Yahoo後綴；市場別穩定，避免某天沒進放量分頁就抓不到。"""
    g = {}
    for ddir in date_dirs:
        files = glob.glob(os.path.join(ddir, "*.xlsx"))
        market, _ = classify(files)
        if market:
            g.update(build_market_map(market))
    return g


def _load_index_raw():
    """讀回既有 index 原始 dict（優先 index.json，其次解析 index.js），失敗回 None。"""
    ij = os.path.join(OUT_DIR, "index.json")
    if os.path.exists(ij):
        try:
            with open(ij, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # 退而求其次：從 index.js 的 __DATAREG 包裝中抽出 JSON
    ijs = os.path.join(OUT_DIR, "index.js")
    if os.path.exists(ijs):
        try:
            with open(ijs, encoding="utf-8") as f:
                t = f.read()
            return json.loads(t[t.index("{"):t.rindex("}") + 1])
        except Exception:
            pass
    return None


def load_existing_index():
    """讀回既有 index → {date: entry}。
    用於合併：原始 data/ 沒有、但 site/data/ 仍有資料的日期不會被洗掉。"""
    by_date = {}
    raw = _load_index_raw()
    if raw:
        for d in raw.get("dates", []):
            if d.get("date"):
                by_date[d["date"]] = d
    return by_date


def load_existing_weekly():
    """讀回既有 weekly_dates → {wkey: entry}（同 load_existing_index 的合併精神）。"""
    by_week = {}
    raw = _load_index_raw()
    if raw:
        for w in raw.get("weekly_dates", []):
            if w.get("wkey"):
                by_week[w["wkey"]] = w
    return by_week


def weekly_json_ok(path):
    if not os.path.exists(path):
        return False
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        return bool(d.get("name"))
    except Exception:
        return False


def weekly_label(wkey):
    """20260529-0604 或 20260529-20260604 → 起訖日 + 顯示字串。"""
    start, end_raw = wkey.split("-")
    end = (start[:4] + end_raw) if len(end_raw) == 4 else end_raw
    label = f"{start[:4]}-{start[4:6]}-{start[6:]} ~ {end[:4]}-{end[4:6]}-{end[6:]}"
    return start, end, label


def reconstruct_weekly_entry(wkey, wdir, global_mkt):
    """從既有 site/data/weekly/<wkey>/ 重建索引項（無原始 Excel 時仍保留週報，
    比照日報「資料夾還在就保留」的合併保護）。"""
    start, end, label = weekly_label(wkey)
    stock_list = []
    for jp in sorted(glob.glob(os.path.join(wdir, "*.json"))):
        base = os.path.basename(jp)
        if base.endswith("_vol.json"):
            continue
        code = base[:-5]
        if not re.fullmatch(r"\d{4}", code):
            continue
        name = ""
        try:
            with open(jp, encoding="utf-8") as f:
                name = json.load(f).get("name", "") or ""
        except Exception:
            pass
        stock_list.append({"code": code, "name": name, "mkt": global_mkt.get(code, "TW")})
    if not stock_list:
        return None
    return {"wkey": wkey, "label": label, "start": start, "end": end, "stocks": stock_list}


def normalize_markets(by_date):
    """市場別穩定：任一天標到 TWO（上櫃）即全部視為 TWO，避免局部重建漏抓。"""
    g = {}
    for entry in by_date.values():
        for s in entry.get("stocks", []):
            if s.get("mkt") == "TWO":
                g[s["code"]] = "TWO"
            g.setdefault(s["code"], s.get("mkt") or "TW")
    for entry in by_date.values():
        for s in entry.get("stocks", []):
            s["mkt"] = g.get(s["code"], "TW")


def main():
    date_dirs = sorted([d for d in glob.glob(os.path.join(DATA_DIR, "*")) if os.path.isdir(d)])
    global_mkt = build_global_market_map(date_dirs)
    # 先載入既有 index：原始 data/ 缺、但 site/data/ 仍在的日期要保留（防止局部重建洗掉舊日期）
    merged = load_existing_index()
    # 代號→名稱後備：BSR-only 的日子（當天沒有放量訊號檔）用既有資料補名稱
    prior_names = {}
    for entry in merged.values():
        for s in entry.get("stocks", []):
            if s.get("name"):
                prior_names.setdefault(s["code"], s["name"])
    built = {}
    for ddir in date_dirs:
        date = os.path.basename(ddir)
        if not re.fullmatch(r"\d{8}", date):
            continue
        label = f"{date[:4]}-{date[4:6]}-{date[6:]}"
        out_day = os.path.join(OUT_DIR, date)
        os.makedirs(out_day, exist_ok=True)
        files = glob.glob(os.path.join(ddir, "*.xlsx")) + glob.glob(os.path.join(ddir, "*_bsr.csv"))
        market, stocks = classify(files)
        print(f"\n=== {date} ===", flush=True)
        market_json = os.path.join(out_day, "market.json")
        if market:
            if SKIP_EXISTING and market_json_ok(market_json):
                print("  market.json (skip)", flush=True)
            else:
                print("  market.json:", process_market(market, market_json), flush=True)
            for tkey in TRACK_SHEETS:
                tj = os.path.join(out_day, f"{tkey}.json")
                if SKIP_EXISTING and market_json_ok(tj):
                    print(f"  {tkey}.json (skip)", flush=True)
                else:
                    n = process_market_track(market, tkey, tj)
                    print(f"  {tkey}.json: {n} 筆", flush=True)
        name_map = None
        stock_list = []
        for code in sorted(stocks):
            stock_json = os.path.join(out_day, f"{code}.json")
            # 名稱優先用 stock_names.json，其次放量訊號檔，再來既有資料，最後留空
            name = NAME_OVERRIDES.get(code)
            if not name:
                if name_map is None:
                    name_map = build_name_map(market)
                name = name_map.get(code) or prior_names.get(code, "")
            mkt = global_mkt.get(code, "TW")  # 未知市場別預設上市
            if SKIP_EXISTING and stock_json_ok(stock_json):
                print(f"  {code} {name}: (skip)", flush=True)
            elif "bsr" in stocks[code]:
                st = process_stock_from_bsr(code, name, stocks[code]["bsr"], stock_json)
                print(f"  {code} {name} (BSR): {st}", flush=True)
            else:
                st = process_stock(code, name, stocks[code], stock_json)
                print(f"  {code} {name}: {st}", flush=True)
            stock_list.append({"code": code, "name": name, "mkt": mkt})
        built[date] = {"date": date, "label": label, "has_market": bool(market), "stocks": stock_list}

    # 合併：本次從原始 data/ 建好的日期為準（覆蓋既有）；其餘既有日期若 site/data/ 仍有資料夾就保留。
    merged.update(built)
    kept = []
    for date in list(merged):
        if date in built or os.path.isdir(os.path.join(OUT_DIR, date)):
            kept.append(date)
        else:
            print(f"  (drop {date}: site/data/{date} 不存在)", flush=True)
            del merged[date]
    normalize_markets(merged)
    index = {"dates": sorted(merged.values(), key=lambda d: d["date"], reverse=True),
             "weekly_dates": []}

    # 代號→名稱：供週報個股命名（週報 Excel 本身常無名稱欄）
    name_by_code = {}
    for entry in merged.values():
        for s in entry.get("stocks", []):
            if s.get("name"):
                name_by_code.setdefault(s["code"], s["name"])

    # ── 週報（比照日報的合併保護：本次建好的為準，其餘既有週若 site/data/weekly/ 仍在則保留）──
    merged_weekly = load_existing_weekly()
    built_weekly = {}
    if os.path.isdir(WEEKLY_DIR):
        week_dirs = sorted([d for d in glob.glob(os.path.join(WEEKLY_DIR, "*")) if os.path.isdir(d)])
        for wdir in week_dirs:
            wkey = os.path.basename(wdir)          # e.g. 20260529-0604
            # 接受 YYYYMMDD-MMDD 或 YYYYMMDD-YYYYMMDD 格式
            if not re.match(r"\d{8}-\d{4,8}$", wkey):
                continue
            start, end, label = weekly_label(wkey)
            out_week = os.path.join(OUT_DIR, "weekly", wkey)
            os.makedirs(out_week, exist_ok=True)
            files = glob.glob(os.path.join(wdir, "*.xlsx"))
            wstocks = classify_weekly(files)
            print(f"\n=== 週報 {wkey} ===", flush=True)
            stock_list = []
            for code in sorted(wstocks):
                name = NAME_OVERRIDES.get(code) or name_by_code.get(code, "")
                w_json = os.path.join(out_week, f"{code}.json")
                if SKIP_EXISTING and weekly_json_ok(w_json):
                    print(f"  {code} {name}: (skip)", flush=True)
                else:
                    st = process_weekly(code, name, wstocks[code], w_json)
                    print(f"  {code} {name} 週報: {st}", flush=True)
                va_json = os.path.join(out_week, f"{code}_vol.json")
                if SKIP_EXISTING and weekly_json_ok(va_json):
                    print(f"  {code} {name} vol: (skip)", flush=True)
                else:
                    vt = process_volume_avg(code, name, wstocks[code], va_json)
                    print(f"  {code} {name} vol: {vt}", flush=True)
                mkt = global_mkt.get(code, "TW")
                stock_list.append({"code": code, "name": name, "mkt": mkt})
            built_weekly[wkey] = {"wkey": wkey, "label": label,
                                  "start": start, "end": end, "stocks": stock_list}

    merged_weekly.update(built_weekly)
    # 既有產出資料夾還在、但本次未重建（無原始 Excel）→ 從資料夾重建索引項，避免漏掉週報
    out_weekly_root = os.path.join(OUT_DIR, "weekly")
    if os.path.isdir(out_weekly_root):
        for wdir in sorted(glob.glob(os.path.join(out_weekly_root, "*"))):
            wkey = os.path.basename(wdir)
            if not os.path.isdir(wdir) or not re.match(r"\d{8}-\d{4,8}$", wkey):
                continue
            if wkey in merged_weekly:
                continue
            entry = reconstruct_weekly_entry(wkey, wdir, global_mkt)
            if entry:
                merged_weekly[wkey] = entry
    # 丟掉資料夾已不存在的週
    for wkey in list(merged_weekly):
        if not os.path.isdir(os.path.join(out_weekly_root, wkey)):
            print(f"  (drop 週報 {wkey}: site/data/weekly/{wkey} 不存在)", flush=True)
            del merged_weekly[wkey]
    index["weekly_dates"] = sorted(merged_weekly.values(), key=lambda d: d["start"], reverse=True)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    built_n = len(built)
    print(f"\nindex.json: {len(index['dates'])} days (本次重建 {built_n} 天，合併保留 {len(index['dates']) - built_n} 天)、"
          f"{len(index['weekly_dates'])} weeks -> {OUT_DIR}", flush=True)

    # ── history_log：找最新日期資料夾裡的 history_log.csv，全量覆蓋輸出 ──
    hist_csv = None
    for ddir in reversed(date_dirs):   # date_dirs 已排序，reversed 取最新
        candidate = os.path.join(ddir, "history_log.csv")
        if os.path.exists(candidate):
            hist_csv = candidate
            break
    if hist_csv:
        hist_json = os.path.join(OUT_DIR, "history_log.json")
        n = process_history_log(hist_csv, hist_json)
        print(f"history_log.json: {n} 筆 (來源: {os.path.basename(os.path.dirname(hist_csv))})", flush=True)
    else:
        print("history_log.csv: 未找到，略過", flush=True)

    emit_js_wrappers()
    print("emitted .js wrappers", flush=True)


if __name__ == "__main__":
    main()
